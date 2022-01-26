import discord
#This imports Discord. Named thing.py because my old bots had their main files in thing.js, and I'm sentimental. 

from datetime import date
from datetime import datetime
from discord.utils import get
from openpyxl import load_workbook
from openpyxl import Workbook

import pytz

est = pytz.timezone('US/Eastern')
utc = pytz.utc

# To run a new month: 
# - Update monthsPossible

wb = Workbook()
wa = Workbook()
wHours = Workbook()
wRegime = Workbook()
wRegimeAverage = Workbook()

testMode = False
if testMode == True:
    messageCount = 100
else:
    messageCount = None


monthPriorTo = pytz.utc.localize(datetime.today()).astimezone(est).month
year = pytz.utc.localize(datetime.today()).astimezone(est).year

secretSecret = False
intents = discord.Intents.all()
client = discord.Client(intents=intents)
monthEndDate = {
    1: 31,
    2: 28,
    3:31,
    4:30,
    5:31,
    6:30,
    7:31,
    8:31,
    9:30,
    10:31,
    11:30,
    12:31
}
@client.event
async def on_ready():
    CADiscord = client.get_guild(523962430179770369)
    serverCreation = pytz.utc.localize(CADiscord.created_at).astimezone(est)
    jonoAbdictation = await CADiscord.get_channel(524453945524092929).fetch_message(713466988428656641)
    endDateJono = pytz.utc.localize(jonoAbdictation.created_at).astimezone(est)
    startDateZach = endDateJono
    zachAbdictation = await CADiscord.get_channel(524454096355459072).fetch_message(845404560125984788)
    endDateZach = endDateJono = pytz.utc.localize(zachAbdictation.created_at).astimezone(est)
    startDateColin = endDateZach
    endDateColin = pytz.utc.localize(datetime(year,monthPriorTo,monthEndDate[monthPriorTo])).astimezone(est)
    jonoRegime = endDateJono-serverCreation
    zachRegime = endDateZach-startDateZach
    colinRegime = endDateColin-startDateColin
    regimes = {"Jono's Regime": endDateJono-serverCreation,"Zach's Regime": endDateZach-startDateZach, "Colin's Regime": endDateColin-startDateColin}
    
    regimeNumToName = {
        0: "Jono's Regime",
        1: "Zach's Regime",
        2: "Colin's Regime"
    }
    
    monthsNumberToWord = {
        1: "January",
        2: "February",
        3: "March",
        4: "April",
        5: "May",
        6: "June",
        7: "July",
        8: "August",
        9: "September",
        10: "October",
        11: "November",
        12: "December"
    }
    monthsPossible = ["December2018"]
    yearsPossible = []
    for number in range(year-(2018+1)):
        yearsPossible.append(2018 + number + 1)
    currentDate = str(monthsNumberToWord[monthPriorTo]) + str(year)
    for yearIn in yearsPossible:
        for month in monthsNumberToWord.values():
            monthsPossible.append(month + str(yearIn))
    for newMonth in range(monthPriorTo-1):
        monthsPossible.append(monthsNumberToWord[newMonth+1] + str(year))
    

    
    validCheckers = []
    validityMessage = await CADiscord.get_channel(524453945524092929).fetch_message(904889501720121366)
    for reaction in validityMessage.reactions:
        async for reactor in reaction.users(limit=None,after=None):
            if not (reactor.id in validCheckers):
                try:
                    bing = CADiscord.get_member(reactor.id).name
                    validCheckers.append(reactor.id)
                except:
                    bing = "incorrect"
    BotRole = CADiscord.get_role(560094358033006629)
    for member in CADiscord.members:
        if BotRole in member.roles and not(member.id in validCheckers):
            validCheckers.append(member.id)
        if secretSecret == True and not(member.id in validCheckers):
            validCheckers.append(member.id)
    
    reactorsList = ""
    for valid in validCheckers:
        reactorsList+=CADiscord.get_member(valid).name
        reactorsList+=", "
    reactorsList = reactorsList[0:len(reactorsList)-2:]
    print(reactorsList)
    checked = []
    totalSorted = 0
    print('We have logged in as {0.user}'.format(client))
    print("\n")
    print("And so it begins.")
    print("Found " + CADiscord.name)
    channelMessagesIndiv = {}
    channelMessagesRoles = {}
    roles = ["Seniors","Juniors","Sophomores","Freshmen","Alumni (2019)","Alumni (2020)", "Alumni (2021)", "Active", "Admins", "Badmin", "Bot", "Member"]
    authors = validCheckers

    channelsCalenderMonthsIndiv = []
    channelsCalenderMonthsTotal = []
    grandTotalAll = {}
    grandTotalTotal = {}
    #channelHoursTotal = {}
    #Finish adding the bits for GrandSheet and also fix the other totals to be for each channel


    regime = []
    
    regimeTotalTotal = []
    regimePeople = []

    timeDict = []
    timeDictTotal = {}
    finalTotalHour = {}
    biggestTotalTimeDictEver = []
    for hour in range(24):
        finalTotalHour[hour] = 0

    for person in validCheckers:
        for month in monthsPossible:
            try:
                grandTotalAll[person][month] = 0
            except:
                grandTotalAll[person] = {month:0}
    for month in monthsPossible:
            grandTotalTotal[month] = 0
    

    for person in validCheckers:
        for hour in range(24):
            try:
                timeDictTotal[person][hour] = 0
            except:
                timeDictTotal[person] = {hour:0}
                
    for ruler in regimes.keys(): 
        regime.append({})
        regimeTotalTotal.append({})
        regimePeople.append({})
    for channel in CADiscord.text_channels:
        print("\nFound #" + channel.name + " (" + str(CADiscord.text_channels.index(channel) + 1) + "/" + str(len(CADiscord.text_channels)) + ")")

        newFile = open("StatsBot\\Channels Caching\\" + channel.name + ".txt", "w", encoding='utf8')

        peopleMessages = {}
        gradeMessages = {}
        number = 0
        channelMonthsIndiv = {}
        channelMonthsTotal = {}
        channelHoursIndiv = {}
        timeHoursIndiv = {}

        messageIDs = []
        messageContents = []
        channelFile = open("StatsBot\\Channels Caching\\" + channel.name + ".txt", "w")
        for hour in range(24):
            timeHoursIndiv[hour] = 0

        async for message in channel.history(limit=messageCount):
            regimeIndiv = {}
            if pytz.utc.localize(message.created_at).astimezone(est).year < year or (pytz.utc.localize(message.created_at).astimezone(est).month < monthPriorTo and pytz.utc.localize(message.created_at).astimezone(est).year == year):
                if len(messageIDs) < 5:
                    messageIDs.append(message.id)
                    messageContents.append(str(message.author.name) + "#" + str(message.author.discriminator) + " [#" + str(message.channel.name) + "]: " + str(message.content) + " (" + str(pytz.utc.localize(message.created_at).astimezone(est).hour) + ":00)")
                if pytz.utc.localize(message.created_at).astimezone(est) <= endDateJono:
                    currentRegime = 0
                if pytz.utc.localize(message.created_at).astimezone(est) > startDateZach and pytz.utc.localize(message.created_at).astimezone(est) <= endDateZach:
                    currentRegime = 1
                if pytz.utc.localize(message.created_at).astimezone(est) > startDateColin and pytz.utc.localize(message.created_at).astimezone(est) <= endDateColin:
                    currentRegime = 2
                if message.author.id in validCheckers:
                    try:
                        peopleMessages[message.author.id]+=1
                    except:
                        peopleMessages[message.author.id] = 1
                    date = pytz.utc.localize(message.created_at).astimezone(est).month
                    dateYear = pytz.utc.localize(message.created_at).astimezone(est).year
                    datePulledTogether = monthsNumberToWord[date] + str(dateYear)
                    try:
                        channelMonthsIndiv[message.author.id][date]+=1
                    except:
                        try:
                            channelMonthsIndiv[message.author.id][date] = 1
                        except:
                            channelMonthsIndiv[message.author.id] = {date:1}
                    try:
                        grandTotalAll[message.author.id][datePulledTogether]+=1
                    except:
                        try:
                            grandTotalAll[message.author.id][datePulledTogether] = 1
                        except:
                            grandTotalAll[message.author.id] = {datePulledTogether:1}
                    try:
                        
                        channelHoursIndiv[message.author.id][pytz.utc.localize(message.created_at).astimezone(est).hour]+=1
                    
                    except:
                        for person in validCheckers:
                            for hour in range(24):
                                try:
                                    channelHoursIndiv[message.author.id][hour] = 0
                                except:
                                    channelHoursIndiv[message.author.id] = {hour:0}
                        channelHoursIndiv[message.author.id][pytz.utc.localize(message.created_at).astimezone(est).hour]+=1
                        
                    try:
                        timeDictTotal[message.author.id][pytz.utc.localize(message.created_at).astimezone(est).hour]+=1
                    except:
                        try:
                            timeDictTotal[message.author.id][pytz.utc.localize(message.created_at).astimezone(est).hour] = 1
                        except:
                            timeDictTotal[message.author.id] = {pytz.utc.localize(message.created_at).astimezone(est).hour:1}

                if channel.name == "rules":
                    print(str(message.author.name) + "#" + str(message.author.discriminator) + " [#" + str(message.channel.name) + "]: " + str(message.content) + " (" + str(pytz.utc.localize(message.created_at).astimezone(est).hour) + ":00)")
                monthCreated = pytz.utc.localize(message.created_at).astimezone(est).month
                yearCreated = pytz.utc.localize(message.created_at).astimezone(est).year
                date = str(monthsNumberToWord[monthCreated]) + str(yearCreated)
                finalTotalHour[pytz.utc.localize(message.created_at).astimezone(est).hour]+=1
                timeHoursIndiv[pytz.utc.localize(message.created_at).astimezone(est).hour]+=1
                try:
                    regimeTotalTotal[currentRegime][channel.name]+=1
                except:
                    regimeTotalTotal[currentRegime][channel.name] = 1
                try:
                    regimePeople[currentRegime][message.author.id]+=1
                except:
                    regimePeople[currentRegime][message.author.id] = 1
                try:
                    channelMonthsTotal[date]+=1
                except:
                    channelMonthsTotal[date] = 1
                try:
                    grandTotalTotal[date]+=1
                except:
                    grandTotalTotal[date] = 1
                try:
                    regime[currentRegime][channel.name][message.author.id] +=1
                except:
                    try:
                        regime[currentRegime][channel.name][message.author.id] = 1
                    except:
                        regime[currentRegime][channel.name] = {message.author.id:1}
                try:
                    member = CADiscord.get_member(message.author.id)
                    for role in member.roles:
                        if role.name in roles:
                            try:
                                gradeMessages[role.id]+=1
                            except:
                                gradeMessages[role.id] = 1
                except:
                    if not(message.author.name in checked):
                        checked.append(message.author.name)
                if number == 0:
                    print(str(message.author.name) + "#" + str(message.author.discriminator) + " [#" + str(message.channel.name) + "]: " + str(message.content))
                
            
            number+=1
            totalSorted+=1
            if number % 500 == 0:
                print(str(number) + " messages sorted in #" + channel.name + " (total: " + str(totalSorted) + ")")
        
        channelFile.write(channel.name)
        for messageIDNum in range(len(messageIDs)):
            try:
                channelFile.write("\n" + str(messageIDs[messageIDNum]) + "|" + str(messageContents[messageIDNum]))
            except:
                try:
                    channelFile.write("\n" + "Invalid ID|" + str(messageContents[messageIDNum]))
                except:
                    try:
                        channelFile.write("\n" + str(messageIDs[messageIDNum]) + "|Invalid Content")
                    except:
                        channelFile.write("\n" + "Invalid ID|Invalid Content")




        channelMessagesIndiv[channel.name] = peopleMessages
        channelMessagesRoles[channel.name] = gradeMessages
        channelsCalenderMonthsIndiv.append(channelMonthsIndiv)
        channelsCalenderMonthsTotal.append(channelMonthsTotal)
        timeDict.append(channelHoursIndiv)
        biggestTotalTimeDictEver.append(timeHoursIndiv)
    
    print(str(biggestTotalTimeDictEver))
    print("Finished reading messages.")

    ws1 = wb.create_sheet("Individuals (ID)")
    ws2 = wb.create_sheet("Individuals (Readable)")
    ws3 = wb.create_sheet("Roles (ID)")
    ws4 = wb.create_sheet("Roles (Readable)")
    listOfChannelsIndiv = list(channelMessagesIndiv.keys())
    listOfChannelsGroup = list(channelMessagesRoles.keys())
    

    infoChannels = ["people-old", "newcomers", "people", "names", "rules", "logging", "announcements", "admin-log", "secret-secret", "chef-sam"]
    boardChannels = ["starboard", "trueboard", "firing-squad", "based-board"]
    spamChannels = ["counting-and-recursion", "bot-spam"]
    HDMFCTInfoChannels = ["historical-death-match-polls", "historical-people-info", "historical-weapons-info", "historical-adjectives-info", "historical-places-info", "historical-contests-info", "historical-contest-specific-info", "fictional-competency-test-polls", "fictional-people-info", "franchises-info", "fictional-adjectives-info", "fictional-minions-info"]
    importantStuffChannels = ["suggestions", "active-users", "admins", "bot-suggestions"]
    generalChannels = ["general", "general-alumni", "programming", "minecraft-bois", "hobbies", "me-me-s", "current-events", "games", "music", "stem", "culture", "foodstuff", "morals-and-ethics", "anime"]
    schoolStuffChannels = ["ca-things", "robotics", "college", "300-seconds", "electives", "homework", "sports"]
    HDMFCTthingsChannels = ["historical-death-match-discussion", "hdm-suggestions", "contest-suggestions", "fct-discussion", "fct-helper-gang"]
    authors.sort()


    nonSpamTotalAuthors = {}
    totalAuthors = {}

    channelTypeToDictAuthors = {
        "Info Channels": {},
        "Boards": {}, 
        "Spam": {}, 
        "HDM/FCT Info": {}, 
        "Important Stuff": {},
        "General": {},
        "School Stuff": {},
        "HDM/FCT Discussions": {}
    }

    channelAuthorDicts = {
        "Info Channels": infoChannels,
        "Boards": boardChannels, 
        "Spam": spamChannels, 
        "HDM/FCT Info": HDMFCTInfoChannels, 
        "Important Stuff": importantStuffChannels,
        "General": generalChannels,
        "School Stuff": schoolStuffChannels,
        "HDM/FCT Discussions": HDMFCTthingsChannels
    }
    listOfChannelTypes = list(channelAuthorDicts.keys())
    for channel in listOfChannelsIndiv:
        channelMessageAuthors = list(channelMessagesIndiv[channel].keys())
        channelMessageAuthors.sort()
        for authorID in channelMessageAuthors:
            number = channelMessagesIndiv[channel][authorID]
            if channel in generalChannels or channel in importantStuffChannels or channel in schoolStuffChannels or channel in HDMFCTthingsChannels:
                try:
                    nonSpamTotalAuthors[authorID]+= number
                except:
                    nonSpamTotalAuthors[authorID] = number   
            for channelType in listOfChannelTypes:  
                if channel in channelAuthorDicts[channelType]:
                    try:
                        channelTypeToDictAuthors[channelType][authorID]+= number
                    except:
                        channelTypeToDictAuthors[channelType][authorID] = number
            try:
                totalAuthors[authorID]+= number
            except:
                totalAuthors[authorID] = number



    
        nonSpamTotalRoles = {}
    totalRoles = {}

    channelTypeToDictRoles = {
        "Info Channels": {},
        "Boards": {}, 
        "Spam": {}, 
        "HDM/FCT Info": {}, 
        "Important Stuff": {},
        "General": {},
        "School Stuff": {},
        "HDM/FCT Discussions": {}
    }

    channelRoleDicts = {
        "Info Channels": infoChannels,
        "Boards": boardChannels, 
        "Spam": spamChannels, 
        "HDM/FCT Info": HDMFCTInfoChannels, 
        "Important Stuff": importantStuffChannels,
        "General": generalChannels,
        "School Stuff": schoolStuffChannels,
        "HDM/FCT Discussions": HDMFCTthingsChannels
    }
    listOfChannelTypesRoles = list(channelRoleDicts.keys())
    for channel in listOfChannelsGroup:
        channelMessageRoles = list(channelMessagesRoles[channel].keys())
        channelMessageRoles.sort()
        for RoleID in channelMessageRoles:
            number = channelMessagesRoles[channel][RoleID]
            if channel in generalChannels or channel in importantStuffChannels or channel in schoolStuffChannels or channel in HDMFCTthingsChannels:
                try:
                    nonSpamTotalRoles[RoleID]+= number
                except:
                    nonSpamTotalRoles[RoleID] = number   
            for channelType in listOfChannelTypes:  
                if channel in channelRoleDicts[channelType]:
                    try:
                        channelTypeToDictRoles[channelType][RoleID]+= number
                    except:
                        channelTypeToDictRoles[channelType][RoleID] = number
            try:
                totalRoles[RoleID]+= number
            except:
                totalRoles[RoleID] = number

    backSortAuthors = {}
    authors = []
    for authorID in totalAuthors.keys():
        value = totalAuthors[authorID]
        try:
            backSortAuthors[value].append(authorID)
        except:
            backSortAuthors[value] = [authorID]

    backSortRoles = {}
    grades = []
    for roleID in totalRoles.keys():
        value = totalRoles[roleID]
        try:
            backSortRoles[value].append(roleID)
        except:
            backSortRoles[value] = [roleID]
    
    

    quantities = list(backSortAuthors.keys())
    quantities.sort()

    quantitiesRoles = list(backSortRoles.keys())
    quantitiesRoles.sort()

    for quanitity in quantitiesRoles:
        for roleID in backSortRoles[quanitity]:
            grades.append(roleID)

    for quantity in quantities:
        for authorID in backSortAuthors[quantity]:
            authors.append(authorID)
    
   
    for channel in listOfChannelsIndiv:
        y = listOfChannelsIndiv.index(channel) + 2
        d3 = ws1.cell(row=1,column=y,value=str(channel))
        d3 = ws2.cell(row=1,column=y,value=str(channel))
        channelMessageAuthors = list(channelMessagesIndiv[channel].keys())
        for quantity in quantities:
            for authorID in backSortAuthors[quantity]:
                try:
                    x = authors.index(authorID) + 2
                    d1 = ws1.cell(row=x,column=1,value=str(authorID))
                    d2 = ws1.cell(row=x,column=y,value=channelMessagesIndiv[channel][authorID])
                    d1 = ws2.cell(row=x,column=1,value=CADiscord.get_member(authorID).name)
                    d2 = ws2.cell(row=x,column=y,value=channelMessagesIndiv[channel][authorID])
                except:
                    zed = 1

    for channel in listOfChannelsGroup:
        y = listOfChannelsGroup.index(channel) + 2
        d3 = ws3.cell(row=1,column=y,value=str(channel))
        d3 = ws4.cell(row=1,column=y,value=str(channel))
        for quantity in quantitiesRoles:
            for roleID in backSortRoles[quantity]:
                try:
                    x = grades.index(roleID) + 2
                    d3 = ws3.cell(row=x,column=1,value=str(roleID))
                    d4 = ws3.cell(row=x,column=y,value=channelMessagesRoles[channel][roleID])
                    d3 = ws4.cell(row=x,column=1,value=CADiscord.get_role(roleID).name)
                    d4 = ws4.cell(row=x,column=y,value=channelMessagesRoles[channel][roleID])
                except:
                    zed = 1
    
    totalOfTotals = [{
        "Non-Spam": {},
        "Info Channels": {},
        "Boards": {}, 
        "Spam": {}, 
        "HDM/FCT Info": {}, 
        "Important Stuff": {},
        "General": {},
        "School Stuff": {},
        "HDM/FCT Discussions": {}
    }, {
        "Non-Spam": {},
        "Info Channels": {},
        "Boards": {}, 
        "Spam": {}, 
        "HDM/FCT Info": {}, 
        "Important Stuff": {},
        "General": {},
        "School Stuff": {},
        "HDM/FCT Discussions": {}
    }]

    grandTotalMembers = 0
    d1 = ws1.cell(row=1,column=len(listOfChannelsIndiv)+2,value="Total")
    d2 = ws2.cell(row=1,column=len(listOfChannelsIndiv)+2,value="Total")
    for authorID in authors:
        x = authors.index(authorID) + 2
        d1 = ws1.cell(row=x,column=len(listOfChannelsIndiv)+2,value=totalAuthors[authorID])
        d2 = ws2.cell(row=x,column=len(listOfChannelsIndiv)+2,value=totalAuthors[authorID])
        grandTotalMembers+=totalAuthors[authorID]
    
    x=0

    y = len(listOfChannelsIndiv)+3
    d1 = ws1.cell(row=1,column=y,value="Non-Spam Total")
    d2 = ws2.cell(row=1,column=y,value="Non-Spam Total")
    for authorID in authors:
        try:
            x = authors.index(authorID) + 2
            d1 = ws1.cell(row=x,column=y,value=nonSpamTotalAuthors[authorID])
            d2 = ws2.cell(row=x,column=y,value=nonSpamTotalAuthors[authorID])
            try:
                totalOfTotals[0]["Non-Spam"]+=nonSpamTotalAuthors[authorID]
            except:
                totalOfTotals[0]["Non-Spam"] = nonSpamTotalAuthors[authorID]
        except:
            x=1
    addition = 4
    print("Channel Types: " + str(listOfChannelTypes))
    for channelType in listOfChannelTypes:
        y = len(listOfChannelsIndiv)+addition
        d1 = ws1.cell(row=1,column=y,value=channelType + " Total")
        d2 = ws2.cell(row=1,column=y,value=channelType + " Total")
        for authorID in authors:
            try:
                x = authors.index(authorID) + 2
                d1 = ws1.cell(row=x,column=y,value=channelTypeToDictAuthors[channelType][authorID])
                d2 = ws2.cell(row=x,column=y,value=channelTypeToDictAuthors[channelType][authorID])
                try:
                    totalOfTotals[0][channelType]+=channelTypeToDictAuthors[channelType][authorID]
                except:
                    totalOfTotals[0][channelType] = channelTypeToDictAuthors[channelType][authorID]
            except:
                b = 0
        addition+=1
        print("Tried " + channelType + " (" + str(addition-4) + "/" + str(len(listOfChannelTypes)) + ")")
     



    grandTotalRoles = 0
    d1 = ws3.cell(row=1,column=len(listOfChannelsGroup)+2,value="Total")
    d2 = ws4.cell(row=1,column=len(listOfChannelsGroup)+2,value="Total")
    for roleID in grades:
        x = grades.index(roleID) + 2
        #print("YESSSSS " + str(totalRoles[roleID]) + "!")
        d1 = ws3.cell(row=x,column=len(listOfChannelsGroup)+2,value=totalRoles[roleID])
        d2 = ws4.cell(row=x,column=len(listOfChannelsGroup)+2,value=totalRoles[roleID])
        grandTotalRoles+=totalRoles[roleID]

    y = len(listOfChannelsGroup)+3
    d1 = ws3.cell(row=1,column=y,value="Non-Spam Total")
    d2 = ws4.cell(row=1,column=y,value="Non-Spam Total")
    
    for roleID in grades:
        try:
            x = grades.index(roleID) + 2
            d1 = ws3.cell(row=x,column=y,value=nonSpamTotalRoles[roleID])
            d2 = ws4.cell(row=x,column=y,value=nonSpamTotalRoles[roleID])
            try:
                totalOfTotals[1]["Non-Spam"]+=nonSpamTotalRoles[roleID]
            except:
                totalOfTotals[1]["Non-Spam"] = nonSpamTotalRoles[roleID]
        except:
            x=1
    addition = 4
    print("Channel Types: " + str(listOfChannelTypes))
    for channelType in listOfChannelTypesRoles:
        y = len(listOfChannelsGroup)+addition
        d1 = ws3.cell(row=1,column=y,value=channelType + " Total")
        d2 = ws4.cell(row=1,column=y,value=channelType + " Total")
        for roleID in grades:
            try:
                x = grades.index(roleID) + 2
                d1 = ws3.cell(row=x,column=y,value=channelTypeToDictRoles[channelType][roleID])
                d2 = ws4.cell(row=x,column=y,value=channelTypeToDictRoles[channelType][roleID])
                try:
                    totalOfTotals[1][channelType]+=channelTypeToDictRoles[channelType][roleID]
                except:
                    totalOfTotals[1][channelType] = channelTypeToDictRoles[channelType][roleID]
            except:
                b = 0
        addition+=1
        print("Tried " + channelType + " (" + str(addition-4) + "/" + str(len(listOfChannelTypes)) + ")")
    
    

    x = len(authors) + 2
    d1 = ws1.cell(row=x,column=1,value="Total")
    d1 = ws2.cell(row=x,column=1,value="Total")
    for channel in channelMessagesIndiv.keys():
        y=list(channelMessagesIndiv.keys()).index(channel) + 2
        channelTotal = 0
        for authorID in channelMessagesIndiv[channel].keys():
            channelTotal+=channelMessagesIndiv[channel][authorID]
        d1 = ws1.cell(row=x,column=y,value=channelTotal)
        d1 = ws2.cell(row=x,column=y,value=channelTotal)
    
    x = len(grades) + 2
    d1 = ws3.cell(row=x,column=1,value="Total")
    d1 = ws4.cell(row=x,column=1,value="Total")
    for channel in channelMessagesRoles.keys():
        y=list(channelMessagesRoles.keys()).index(channel) + 2
        channelTotal = 0
        for roleID in channelMessagesRoles[channel].keys():
            channelTotal+=channelMessagesRoles[channel][roleID]
        d1 = ws3.cell(row=x,column=y,value=channelTotal)
        d1 = ws4.cell(row=x,column=y,value=channelTotal)

    x = len(authors) + 2
    y = len(channelMessagesIndiv.keys())+2
    d1 = ws1.cell(row=x,column=y,value=grandTotalMembers)
    d1 = ws2.cell(row=x,column=y,value=grandTotalMembers)

    x=len(grades)+2
    y=len(channelMessagesRoles.keys())+2
    d1 = ws3.cell(row=x,column=y,value=grandTotalRoles)
    d1 = ws4.cell(row=x,column=y,value=grandTotalRoles)

    additionIndiv = 3
    for totalTotalChannel in totalOfTotals[0].keys():
        x = len(authors) + 2
        y = len(channelMessagesIndiv.keys())+additionIndiv
        d1 = ws1.cell(row=x,column=y,value=totalOfTotals[0][totalTotalChannel])
        d1 = ws2.cell(row=x,column=y,value=totalOfTotals[0][totalTotalChannel])
        additionIndiv+=1
    
    additionGroup = 3
    for totalTotalChannel in totalOfTotals[1].keys():
        x = len(grades)+2
        y = len(channelMessagesRoles.keys())+additionGroup
        d1 = ws3.cell(row=x,column=y,value=totalOfTotals[1][totalTotalChannel])
        d1 = ws4.cell(row=x,column=y,value=totalOfTotals[1][totalTotalChannel])
        additionGroup+=1

    print("\n\n")
    userMeetsJoinTime = {}
    newcomers = CADiscord.get_channel(524456819515719700)
    for member in validCheckers:
        joinTime = pytz.utc.localize(CADiscord.get_member(member).joined_at).astimezone(est)
        userMeetsJoinTime[CADiscord.get_member(member).id] = joinTime
    async for message in newcomers.history(limit=None):
        try:
            if pytz.utc.localize(message.created_at).astimezone(est) < userMeetsJoinTime[message.author.id]:
                print("Overwriting ~" + message.author.name + "~ " + str(joinTime) + " with " + str(pytz.utc.localize(message.created_at).astimezone(est)))
                joinTime = pytz.utc.localize(message.created_at).astimezone(est)
                userMeetsJoinTime[message.author.id] = joinTime
        except:
            zef = False

    print("\n")
    daysOfFuturePast = {}
    d1 = ws1.cell(row=1,column=len(listOfChannelsIndiv)+additionIndiv,value="Date")
    d2 = ws2.cell(row=1,column=len(listOfChannelsIndiv)+additionIndiv,value="Date")
    for authorID in authors:
        x = authors.index(authorID) + 2
        
        todayDate = pytz.utc.localize(datetime(year,monthPriorTo,monthEndDate[monthPriorTo])).astimezone(est)
        user = client.get_user(authorID)
        try:
            user = CADiscord.get_member(authorID)
            joinTime = userMeetsJoinTime[authorID]
            print("Date for " + str(user.name) + ": " + str(joinTime))
            channel = CADiscord.get_channel(524456819515719700)
                
            dateOnServer = todayDate-joinTime
            daysOn = int(str(dateOnServer).split(" days, ")[0])
            d1 = ws1.cell(row=x,column=len(listOfChannelsIndiv)+additionIndiv,value=daysOn)
            d2 = ws2.cell(row=x,column=len(listOfChannelsIndiv)+additionIndiv,value=daysOn)
            daysOfFuturePast[authorID] = daysOn
        except:
            try: 
                print("Couldn't find " + str(authorID) + " (" + str(CADiscord.get_member(authorID).name) + ")")
            except:
                print("Couldn't find " + str(authorID))

    print("\n\n")
    additionIndiv+=1
    d1 = ws1.cell(row=1,column=len(listOfChannelsIndiv)+additionIndiv,value="Message Average Total")
    d2 = ws2.cell(row=1,column=len(listOfChannelsIndiv)+additionIndiv,value="Message Average Total")
    d1 = ws1.cell(row=1,column=len(listOfChannelsIndiv)+additionIndiv+1,value="Message Average No Spam")
    d2 = ws2.cell(row=1,column=len(listOfChannelsIndiv)+additionIndiv+1,value="Message Average No Spam")
    for authorID in daysOfFuturePast.keys():
        x = authors.index(authorID) + 2
        totalValue = totalAuthors[authorID]
        d1 = ws1.cell(row=x,column=len(listOfChannelsIndiv)+additionIndiv,value=totalValue/daysOfFuturePast[authorID])
        d2 = ws2.cell(row=x,column=len(listOfChannelsIndiv)+additionIndiv,value=totalValue/daysOfFuturePast[authorID])
        try:
            noSpamValue = nonSpamTotalAuthors[authorID]
            d1 = ws1.cell(row=x,column=len(listOfChannelsIndiv)+additionIndiv+1,value=noSpamValue/daysOfFuturePast[authorID])
            d2 = ws2.cell(row=x,column=len(listOfChannelsIndiv)+additionIndiv+1,value=noSpamValue/daysOfFuturePast[authorID])
        except:
            b=0
    
    daysPerChannel = {}
    maxRow = len(validCheckers) + 2
    d1 = ws1.cell(row=maxRow,column=1,value="Days Existing")
    d1 = ws2.cell(row=maxRow,column=1,value="Days Existing")
    for channel in listOfChannelsIndiv:
        todayDate = datetime(year,monthPriorTo,monthEndDate[monthPriorTo])
        dateOnServer = todayDate-discord.utils.get(CADiscord.channels, name=channel).created_at
        daysOn = int(str(dateOnServer).split(" days, ")[0])
        daysPerChannel[channel] = daysOn
        d1 = ws1.cell(row=maxRow, column=listOfChannelsIndiv.index(channel)+2,value=daysOn)
        d1 = ws2.cell(row=maxRow, column=listOfChannelsIndiv.index(channel)+2,value=daysOn)

    x = maxRow + 1
    d1 = ws1.cell(row=x,column=1,value="Average Messages")
    d1 = ws2.cell(row=x,column=1,value="Average Messages")
    for channel in channelMessagesIndiv.keys():
        y=list(channelMessagesIndiv.keys()).index(channel) + 2
        channelTotal = 0
        for authorID in channelMessagesIndiv[channel].keys():
            channelTotal+=channelMessagesIndiv[channel][authorID]
        d1 = ws1.cell(row=x,column=y,value=channelTotal/daysPerChannel[channel])
        d1 = ws2.cell(row=x,column=y,value=channelTotal/daysPerChannel[channel])


    channelSheetsHours = {}
    channelSheetsMonths = {}
    for channel in listOfChannelsIndiv:
        if len(channel) > 29:
            channelSheetsMonths[channel] = wa.create_sheet("#" + channel[0:29:])
            print("Created Sheet: " + "#" + channel[0:29:])
            channelSheetsHours[channel] = wHours.create_sheet("#" + channel[0:29:])
        else:
            channelSheetsMonths[channel] = wa.create_sheet("#" + channel)
            channelSheetsHours[channel] = wHours.create_sheet("#" + channel)
            print("Created Sheet: " + "#" + channel)


    channelEarlyMonth = {}
    for channel in listOfChannelsIndiv:
        channelObject = discord.utils.get(CADiscord.channels, name=channel)
        startMonth = channelObject.created_at.month + 1
        yearAddition = 0
        if startMonth == 13:
            startMonth = 1
            yearAddition = 1
        date = monthsNumberToWord[startMonth] + str(channelObject.created_at.year + yearAddition)
        channelEarlyMonth[channel] = date
    
    i = 0
    for channelDict in channelsCalenderMonthsIndiv:
        channelName = listOfChannelsIndiv[i]
        ws = channelSheetsMonths[channelName]
        validYet = False
        columnNum = 2
        for month in monthsPossible:
            if month == channelEarlyMonth[channelName]:
                validYet = True
            if validYet == True:
                d1 = ws.cell(row=1, column=columnNum,value=month)
                columnNum += 1
         
        i+=1

    i = 0
    for channelDict in channelsCalenderMonthsIndiv:
        channelName = listOfChannelsIndiv[i]
        ws = channelSheetsMonths[channelName]
        for person in channelDict.keys():
            ws.cell(row=list(channelDict.keys()).index(person)+2,column=1,value=CADiscord.get_member(person).name)
            personDict = channelDict[person]
            for cellIndex in range(2, len(monthsPossible)-monthsPossible.index(channelEarlyMonth[channelName])+2):
                if ws.cell(row=1,column=cellIndex).value in list(personDict.keys()):
                    ws.cell(row=list(channelDict.keys()).index(person)+2,column=cellIndex,value=personDict[ws.cell(row=1,column=cellIndex).value])
                    
        i+=1

    i = 0

    for channelDict in channelsCalenderMonthsTotal:

        channelName = listOfChannelsIndiv[i]
        channelDictPeople = channelsCalenderMonthsIndiv[i]
        ws = channelSheetsMonths[channelName]
        ws.cell(row=len(list(channelDictPeople.keys()))+2,column=1,value="Total #" + channelName)
        grandTotal = 0
        if channelEarlyMonth[channelName] != currentDate:
            for cellIndex in range(2, len(monthsPossible)-monthsPossible.index(channelEarlyMonth[channelName])+2):
                if ws.cell(row=1,column=cellIndex).value in list(channelDict.keys()):
                    ws.cell(row=len(list(channelDictPeople.keys()))+2,column=cellIndex,value=channelDict[ws.cell(row=1,column=cellIndex).value])
                    grandTotal+=channelDict[ws.cell(row=1,column=cellIndex).value]

            i+=1
            ws.cell(row=1,column=len(monthsPossible)-monthsPossible.index(channelEarlyMonth[channelName])+2,value="Total")

            ws.cell(row=len(list(channelDictPeople.keys()))+2,column=len(monthsPossible)-monthsPossible.index(channelEarlyMonth[channelName])+2,value=grandTotal)
    
    trueSheet1 = wa.create_sheet("Grand Total", 0)
    i = 0
    for monthNum in range(2,len(monthsPossible)+2):
        trueSheet1.cell(row=1,column=monthNum,value=monthsPossible[monthNum-2])
   
    for person in grandTotalAll.keys():
        trueSheet1.cell(row=i+2,column=1,value=CADiscord.get_member(person).name)
        columnNum = 2
        for month in grandTotalAll[person].keys():
            #Dec. 2018 = 2
            #print(month + "/" + str(columnNum))
            trueSheet1.cell(row=i+2,column=columnNum,value=grandTotalAll[person][month])
            if columnNum > 38:
                print("printing " + str(grandTotalAll[person][month]) + " at (" + str(i+2) + "," + str(columnNum) + ")")
                print("\t" + str(grandTotalAll[person].keys()))
            columnNum+=1
        i+=1
    trueSheet1.cell(row=i+2,column=1,value="Grand Total")
    columnNum = 2
    for month in grandTotalTotal:
        trueSheet1.cell(row=i+2,column=columnNum,value=grandTotalTotal[month])

        columnNum+=1


    i = 2
    trueSheet2 = wHours.create_sheet("Grand Total",0)

    for hour in range(2,26):
        trueSheet2.cell(row=1,column=hour,value=str(hour-2)+":00")
    for person in timeDictTotal.keys():
        try:
            trueSheet2.cell(row=i,column=1,value=CADiscord.get_member(person).name)
            columnNum = 2
            for hour in timeDictTotal[person].keys():
                trueSheet2.cell(row=i,column=columnNum,value=timeDictTotal[person][hour])
                columnNum+=1
            i+=1
        except:
            print("Could not find " + str(person))
    trueSheet2.cell(row=i,column=1,value="Total")
    columnNum = 2
    for hour in finalTotalHour.keys():
        trueSheet2.cell(row=i,column=columnNum,value=finalTotalHour[hour])
        columnNum+=1
    foo = 0
    for channel in listOfChannelsIndiv:
        channelDict = timeDict[foo]
        i = 2
        sheet = channelSheetsHours[channel]
        for hour in range(2,26):
            sheet.cell(row=1,column=hour,value=str(hour-2)+":00")
        for person in channelDict.keys():
            try:
                sheet.cell(row=i,column=1,value=CADiscord.get_member(person).name)
                columnNum = 2
                for hour in channelDict[person].keys():
                    sheet.cell(row=i,column=columnNum,value=channelDict[person][hour])
                    columnNum+=1
                i+=1
            except:
                print("Could not find " + str(person))
        newTimeDictTotal = biggestTotalTimeDictEver[foo]
        columnNum = 1
        sheet.cell(row=len(channelDict.keys())+2,column=columnNum,value="Total")
        columnNum+=1
        for hour in newTimeDictTotal.keys():
            sheet.cell(row=i,column=columnNum,value=newTimeDictTotal[hour])
            columnNum+=1
        foo+=1
    
    trueSheet3 = wRegime.create_sheet("Grand Total",0)
    regimeNum = 0
    fullNum = 0
    i = 2
    for person in validCheckers:
        trueSheet3.cell(row=i+1,column=1,value=CADiscord.get_member(person).name)
        i+=1
    regimeNumber = 0
    
    for regimeDict in regime:
        i = 0
        trueSheet3.cell(row=1,column=2 + fullNum,value=str(list(regimes.keys())[regimeNum]))
        for channel in regimeDict.keys():
            trueSheet3.cell(row=2,column=2+fullNum + i,value=channel)
            for person in regimeDict[channel].keys():
                for personNum in range(len(validCheckers)):
                    try:
                        if trueSheet3.cell(row=personNum+3,column=1).value == CADiscord.get_member(person).name:
                            trueSheet3.cell(row=personNum+3,column=2+fullNum + i,value=regimeDict[channel][person])
                    except:
                        beep = False
            trueSheet3.cell(row=len(validCheckers)+3,column=2+fullNum+i,value=regimeTotalTotal[regimeNumber][channel])
            len(validCheckers)-1
            i+=1
        regimeNum+=1
        regimeNumber+=1
        fullNum+=len(regimeDict.keys())
    trueSheet3.cell(row=len(validCheckers)+3,column=1,value="Total")

        
    trueSheet = wRegimeAverage.create_sheet("Grand Total",0)
    regimeNum = 0
    fullNum = 0
    i = 2
    for person in validCheckers:
        trueSheet.cell(row=i+1,column=1,value=CADiscord.get_member(person).name)
        i+=1
    regimeNumber = 0
    
    for regimeDict in regime:
        daysOfRegime = list(regimes.values())[regimeNumber].days
        i = 0
        trueSheet.cell(row=1,column=2 + fullNum,value=str(list(regimes.keys())[regimeNum]))
        for channel in regimeDict.keys():
            trueSheet.cell(row=2,column=2+fullNum + i,value=channel)
            for person in regimeDict[channel].keys():
                for personNum in range(len(validCheckers)):
                    try:
                        if trueSheet.cell(row=personNum+3,column=1).value == CADiscord.get_member(person).name:
                            trueSheet.cell(row=personNum+3,column=2+fullNum + i,value=regimeDict[channel][person]/daysOfRegime)
                    except:
                        beep = False
            trueSheet.cell(row=len(validCheckers)+3,column=2+fullNum+i,value=regimeTotalTotal[regimeNumber][channel]/daysOfRegime)
            len(validCheckers)-1
            i+=1
        regimeNum+=1
        regimeNumber+=1
        fullNum+=len(regimeDict.keys())
    trueSheet.cell(row=len(validCheckers)+3,column=1,value="Total")

        
    wRegimeAverage.save("StatBotRegimesAverages" + str(monthsNumberToWord[monthPriorTo]) + str(year) + ".xlsx")
    wRegime.save("StatBotRegimes" + str(monthsNumberToWord[monthPriorTo]) + str(year) + ".xlsx")
    wa.save("StatBotCalender" + str(monthsNumberToWord[monthPriorTo]) + str(year) + ".xlsx")
    wHours.save("StatBotClock" + str(monthsNumberToWord[monthPriorTo]) + str(year) + ".xlsx")

    print("Saved as " + str(monthsNumberToWord[monthPriorTo]) + str(year))
    
    
    wb.save("StatBotInfo" + str(monthsNumberToWord[monthPriorTo]) + str(year) + ".xlsx")
    print("Completed!")



client.run('NjYyNzg4MTk1NzM3NDAzNDQy.Xg_Dmw.dyg4Kch4KxX6C6bDZAcx-Le2TVs')