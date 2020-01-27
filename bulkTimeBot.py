import discord
#This imports Discord. Named thing.py because my old bots had their main files in thing.js, and I'm sentimental. 
import xlrd
import xlwt
import pytz
from xlwt import Workbook
from xlutils.copy import copy 
from xlrd import open_workbook
from datetime import date
from datetime import datetime
from discord.utils import get
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from PIL import Image
import numpy as np

wb = Workbook()

est = pytz.timezone('US/Eastern')
utc = pytz.utc
fmt = '%Y-%m-%d %H:%M:%S %Z%z'

client = discord.Client()
def error():
    embed = discord.Embed(title="Error", description="Sorry, something went wrong.", color=0xFF9900)
    embed.set_footer(text="Created by The Invisible Man", icon_url="https://cdn.discordapp.com/avatars/366709133195476992/01cb7c2c7f2007d8b060e084ea4eb6fd.png?size=512")
    return embed
def findUser(userID):
    return user
@client.event
async def on_ready():
    print('We have logged in as {0.user}'.format(client))
    print("\n")

@client.event
async def on_message(message):
    if message.author == client.user:
        return
    if message.content.startswith("&&I bid y'all adieu") and message.author.id == 366709133195476992:
        time.sleep(2)
        await message.channel.send("Farewell, and have a good night!")
        time.sleep(2)
        await message.guild.leave()
    if message.content.startswith("&&help"):
        print("Help request made by " + message.author.name + " at " + str(message.created_at) + " on guild " + message.guild.name)
        embed = discord.Embed(title="Help", description="Hello, I'm BulkTimeStatBot! Here are some of my functions: ", color=0xFF9900)
        embed.set_footer(text="Created by The Invisible Man", icon_url="https://cdn.discordapp.com/avatars/366709133195476992/01cb7c2c7f2007d8b060e084ea4eb6fd.png?size=512")
        await message.channel.send(embed=embed)
    if message.content.startswith('&&statCountAllServer') and message.author.id == 366709133195476992:
        CoolPeeps = client.get_guild(364116594324013087)
        Starserver = client.get_guild(385515066188890123)
        TheCADiscord = client.get_guild(523962430179770369)
        Level99 = client.get_guild(620758472451162142)
        testserver = client.get_guild(620964009247768586)
        print("Yes: " + CoolPeeps.name)
        guildList = [CoolPeeps, Starserver, TheCADiscord, Level99, testserver]
        for serverActive in guildList:
            print(serverActive.name)
            wb = Workbook()
            sheet1 = wb.add_sheet('Sheet 1')
            i = 1
            zed = 0
            memberList = serverActive.members
            memberQuant = len(memberList)
            for member in memberList:
                sheet1.write(i, 0, str(member.id))
                i+=1
                print("Added " + member.name + "(" + str(member.id) + ")")
            channelList = serverActive.text_channels
            channelQuant = len(channelList)
            x = 1
            for channel in channelList:
                if channel.name != "robot-game" and channel.name != "starfall-private-space" and channel.name != "ca-nerd-squad":
                    sheet1.write(0, x, channel.name)
                    x+=1
                    print("Added #" + channel.name)
            wb.save(serverActive.name + ".xls")
            server = serverActive.text_channels       
            for channel in server:
                if channel.name != "robot-game" and channel.name != "starfall-private-space" and channel.name != "ca-nerd-squad" and channel.name != "vent":
                    y = 1
                    authorMessageQuant = {
                    }
                    for member in memberList:
                        authorMessageQuant[str(member.id)] = 0
                    async for message in channel.history(limit=None):
                        for member in memberList:
                            if member.name == message.author.name:
                                #print("Found author!")
                                authorMessageQuant[str(member.id)] = str(int(authorMessageQuant[str(member.id)])+ 1)
                    wr = xlrd.open_workbook("C:\\Users\\Sebastian_Polge\\OneDrive-CaryAcademy\\Documents\\meNewBot\\Verity\\StatsBot\\" + serverActive.name + ".xls") 
                    sheet = wr.sheet_by_index(0) 
                    for stuff in range(sheet.ncols):
                        if sheet.cell_value(0,stuff) == channel.name:
                            print("Channel is found!") 
                            y = stuff
                    for member in authorMessageQuant:
                        x = 1
                        for mStuff in range(sheet.nrows):
                            if sheet.cell_value(mStuff,0) == member:
                                print("Author is found!" + sheet.cell_value(mStuff, 0) + "/" + member) 
                                x = mStuff
                        print(str(x) + "/" + str(memberQuant) + ";" + str(y) + "/" + str(channelQuant) + " #" + str(channel.name))
                        messageCount = authorMessageQuant[member]
                        sheet1.write(x, y, int(messageCount))
                        zed = y + 1
                        wb.save(message.guild.name + ".xls")
            x = 1
            sheet1.write(0, zed, "Message Total:")
            sheet1.write(0, zed + 1, "Date Joined:")
            sheet1.write(0, zed + 2, "Days on Server:")
            sheet1.write(0, zed + 3, "Message Average:")
            for member in memberList:
                string = ""
                n = y
                while n > 0:
                    n, remainder = divmod(n - 1, 26)
                    string = chr(65 + remainder) + string
                print(string)
                value1 = "B" + str(x + 1)
                value2 = string + str(x + 1)
                sheet1.write(x, zed, xlwt.Formula("SUM(" + value1 + ":" + value2 + ")"))
                print("Zed: " + str(zed) + ", x: " + str(x))
                dateTimeFull = member.joined_at
                dateTimeSplit = str(dateTimeFull).split()
                sheet1.write(x, zed + 1, dateTimeSplit[0])
                print("Zed: " + str(zed) + ", x: " + str(x))
                dateTimeFull = member.joined_at
                dateJoined = datetime.date(dateTimeFull)
                dateTimeSplit = str(dateTimeFull).split()
                currentDate = datetime.date(datetime.now())
                delta = currentDate - dateJoined
                sheet1.write(x, zed + 2, delta.days)
                sheet1.write(x, zed + 3, xlwt.Formula("SUM(" + value1 + ":" + value2 + ")/" + str(delta.days))) 
                print("Zed: " + str(zed) + ", x: " + str(x))
                x+=1
            wb.save(serverActive.name + ".xls")
            print("Complete!")
    if message.content.startswith('&&updateComplete') and message.author.id == 366709133195476992:
        for serverActive in client.guilds:
            rb = xlrd.open_workbook("C:\\Users\\Sebastian_Polge\\OneDrive-CaryAcademy\\Documents\\meNewBot\\Verity\\StatsBot\\Complete.xls") 
            serverStatBook = xlrd.open_workbook("C:\\Users\\Sebastian_Polge\\OneDrive-CaryAcademy\\Documents\\meNewBot\\Verity\\StatsBot\\" + serverActive.name + ".xls") 
            serverStat = serverStatBook.sheet_by_index(0)
            sheetOne = wc.add_sheet(serverActive.name)
            for i in range(serverStat.ncols):
                for j in range(serverStat.nrows):
                    if i != serverStat.ncols - 1 and i != serverStat.ncols - 4:
                        sheetOne.write(j, i, serverStat.cell_value(j,i))
            for j in range(serverStat.nrows):
                if j != 0:
                    string = ""
                    n = serverStat.ncols - 5
                    while n > 0:
                        n, remainder = divmod(n - 1, 26)
                        string = chr(65 + remainder) + string
                    value1 = "B" + str(j + 1)
                    value2 = string + str(j + 1)
                    formula = "SUM(" + value1 + ":" + value2 + ")"
                    sheetOne.write(j, serverStat.ncols - 4, xlwt.Formula(formula))
                    print("Formula: " + formula)
                    print("name: " + serverStat.cell_value(j,0))
                    member = serverActive.get_member(int(serverStat.cell_value(j,0)))
                    dateTimeFull = member.joined_at
                    dateJoined = datetime.date(dateTimeFull)
                    dateTimeSplit = str(dateTimeFull).split()
                    currentDate = datetime.date(datetime.now())
                    delta = currentDate - dateJoined
                    sheetOne.write(j, serverStat.ncols - 1, xlwt.Formula("SUM(" + value1 + ":" + value2 + ")/" + str(delta.days)))
                    wc.save("Complete.xls")
        print("Completed.")
    if message.content.startswith('&&timeCountAll'):
        memberList = []
        for i in client.guilds: 
            print(i.id)
            for person in i.members:
                if person not in memberList:
                    memberList.append(person)
        print('Members Finished')
        number = 1
        wc = Workbook()
        wz = Workbook()

        sheet1 = wc.add_sheet('Sheet 1')
        sheet2 = wz.add_sheet('Sheet 1')

        sheet1.write(0, 1, "0:00")
        sheet1.write(0, 2, "1:00")
        sheet1.write(0, 3, "2:00")
        sheet1.write(0, 4, "3:00")
        sheet1.write(0, 5, "4:00")
        sheet1.write(0, 6, "5:00")
        sheet1.write(0, 7, "6:00")
        sheet1.write(0, 8, "7:00")
        sheet1.write(0, 9, "8:00")
        sheet1.write(0, 10, "9:00")
        sheet1.write(0, 11, "10:00")
        sheet1.write(0, 12, "11:00")
        sheet1.write(0, 13, "12:00")
        sheet1.write(0, 14, "13:00")
        sheet1.write(0, 15, "14:00")
        sheet1.write(0, 16, "15:00")
        sheet1.write(0, 17, "16:00")
        sheet1.write(0, 18, "17:00")
        sheet1.write(0, 19, "18:00")
        sheet1.write(0, 20, "19:00")
        sheet1.write(0, 21, "20:00")
        sheet1.write(0, 22, "21:00")
        sheet1.write(0, 23, "22:00")
        sheet1.write(0, 24, "23:00")

        sheet2.write(0, 1, "0:00")
        sheet2.write(0, 2, "1:00")
        sheet2.write(0, 3, "2:00")
        sheet2.write(0, 4, "3:00")
        sheet2.write(0, 5, "4:00")
        sheet2.write(0, 6, "5:00")
        sheet2.write(0, 7, "6:00")
        sheet2.write(0, 8, "7:00")
        sheet2.write(0, 9, "8:00")
        sheet2.write(0, 10, "9:00")
        sheet2.write(0, 11, "10:00")
        sheet2.write(0, 12, "11:00")
        sheet2.write(0, 13, "12:00")
        sheet2.write(0, 14, "13:00")
        sheet2.write(0, 15, "14:00")
        sheet2.write(0, 16, "15:00")
        sheet2.write(0, 17, "16:00")
        sheet2.write(0, 18, "17:00")
        sheet2.write(0, 19, "18:00")
        sheet2.write(0, 20, "19:00")
        sheet2.write(0, 21, "20:00")
        sheet2.write(0, 22, "21:00")
        sheet2.write(0, 23, "22:00")
        sheet2.write(0, 24, "23:00")

        wc.save("TimesMan.xls")
        wz.save("TimesManName.xls")
        authorDict = {}
        for beep in memberList:
            print(beep.name + ": " + str(beep.id))
            sheet1.write(number, 0, str(beep.id))
            sheet2.write(number, 0, str(beep.name))
            number+=1
        wc.save("TimesMan.xls")
        wz.save("TimesManName.xls")
        timesTW = []
        
        for z in range(0, 24):
            newDict = {}
            for beep in memberList:
                newDict[beep.name] = 0
            timesTW.append(newDict)
        for i in client.guilds:
            if i.name != "The CA Discord" and i.name != "Cool Peeps":
                for chan in i.text_channels:   
                    if chan.name != "robot-game" and chan.name != "starfall-private-space" and chan.name != "ca-nerd-squad":
                        print(i.name + ": " + chan.name) 
                        async for mess in chan.history(limit=None):
                            if mess.author in memberList:
                                messageTimeUTC = mess.created_at
                                source_time_zone = pytz.timezone("Etc/UTC")
                                source_date_with_timezone = source_time_zone.localize(messageTimeUTC)
                                target_time_zone = est
                                target_date_with_timezone = source_date_with_timezone.astimezone(target_time_zone)
                                hourEST = int(target_date_with_timezone.hour)
                                correctDict = timesTW[hourEST - 1]
                                correctDict[mess.author.name] +=1
        for i in range(len(timesTW)):
            for m in range(len(timesTW[i])):
                rowNum = m + 1
                curDict = list(timesTW[i].values())
                print("Value of Col " + str(i + 1) + ":" + str(curDict[m]))
                sheet1.write(rowNum, i + 1, curDict[m])
                sheet2.write(rowNum, i + 1, curDict[m])

                
        wc.save("TimesMan.xls")
        wz.save("TimesManName.xls")
    





            

                    

        



            

               

                
client.run('NjcxMzU0NjYxNDYyMjEyNjI5.Xi7uLQ.IGEc9Mtrihxac5SwdVITp2PvBwI')