import discord

from datetime import date
from datetime import datetime
from discord.utils import get
from openpyxl import load_workbook

from openpyxl import Workbook
wb = Workbook()
wb = load_workbook(filename = 'Statculus.xlsx')

wm = Workbook()

grades = ["Freshmen", "Sophmore", "Junior", "Senior"]

gradeResponses = {"Freshmen": 0, "Sophmore": 0, "Junior": 0, "Senior": 0,}
gradeMonsters = [{}, {}, {}, {}]
kaijuChoices = {"Mokele-Mbembe": {"Freshmen": 0, "Sophmore": 0, "Junior": 0, "Senior": 0}, "Destroyah": {"Freshmen": 0, "Sophmore": 0, "Junior": 0, "Senior": 0}, "Bunyip": {"Freshmen": 0, "Sophmore": 0, "Junior": 0, "Senior": 0}, "Giant Condor": {"Freshmen": 0, "Sophmore": 0, "Junior": 0, "Senior": 0}}



#kaijuLovecraft = {"Mokele-Mbembe": {"Azathoth": 0, "Nyarlathotep": 0, "Shoggoth": 0, "The Colour Out of Space": 0}, "Destroyah": {"Azathoth": 0, "Nyarlathotep": 0, "Shoggoth": 0, "The Colour Out of Space": 0}, "Bunyip": {"Azathoth": 0, "Nyarlathotep": 0, "Shoggoth": 0, "The Colour Out of Space": 0}, "Giant Condor": {"Azathoth": 0, "Nyarlathotep": 0, "Shoggoth": 0, "The Colour Out of Space": 0}}
kaijuLovecraft = {"Azathoth": {"Mokele-Mbembe": 0, "Destroyah": 0, "Bunyip": 0, "Giant Condor": 0}, "Nyarlathotep": {"Mokele-Mbembe": 0, "Destroyah": 0, "Bunyip": 0, "Giant Condor": 0}, "Shoggoth": {"Mokele-Mbembe": 0, "Destroyah": 0, "Bunyip": 0, "Giant Condor": 0}, "The Colour Out of Space": {"Mokele-Mbembe": 0, "Destroyah": 0, "Bunyip": 0, "Giant Condor": 0}}



gradeLovecrafts = [{}, {}, {}, {}]
lovecraftChoices = {"Azathoth": {"Freshmen": 0, "Sophmore": 0, "Junior": 0, "Senior": 0}, "Nyarlathotep": {"Freshmen": 0, "Sophmore": 0, "Junior": 0, "Senior": 0}, "Shoggoth": {"Freshmen": 0, "Sophmore": 0, "Junior": 0, "Senior": 0}, "The Colour Out of Space": {"Freshmen": 0, "Sophmore": 0, "Junior": 0, "Senior": 0}}

sheet_ranges = wb['Sheet1']
responseNum = 0

for row in sheet_ranges.rows:
    responseNum+=1

for response in range(2, responseNum+1):
    ID = str(sheet_ranges.cell(row=response, column=1).value)
    kaiju = str(sheet_ranges.cell(row=response, column=14).value)
    grade = str(sheet_ranges.cell(row=response, column=8).value)
    lovecraftianHorror = str(sheet_ranges.cell(row=response, column=11).value)
    #print(ID + ": " + grade + "/" + kaiju + "/" + lovecraftianHorror)
    try:
        kaijuChoices[kaiju][grade]+=1
    except:
        kaijuChoices[kaiju][grade] = 0
    try:
        lovecraftChoices[lovecraftianHorror][grade]+=1
    except:
        lovecraftChoices[lovecraftianHorror][grade] = 0
    try:
        kaijuLovecraft[lovecraftianHorror][kaiju]+=1
    except:
        kaijuLovecraft[lovecraftianHorror][kaiju] = 0

resultsFile = open("Kaiju Results.txt", "w", encoding='utf8')
for kaiju in kaijuChoices.keys():
    kaijuString = kaiju + ":"
    grandTotal = 0
    for grade in kaijuChoices[kaiju].keys():
        kaijuString = kaijuString + "\n   -" + grade + ": " + str(kaijuChoices[kaiju][grade])
        grandTotal+=kaijuChoices[kaiju][grade]
    kaijuString = kaijuString + "\n   -Total: " + str(grandTotal) + "/" + str(responseNum-1) + " (" + str(round(grandTotal/(responseNum-1)*100,2)) + "%)"
    resultsFile.write(kaijuString + "\n")
resultsFile.close()

resultsFile = open("Lovecraftian Horror Results.txt", "w", encoding='utf8')
for lovecraft in lovecraftChoices.keys():
    lovecraftString = lovecraft + ":"
    grandTotal = 0
    for grade in lovecraftChoices[lovecraft].keys():
        lovecraftString = lovecraftString + "\n   -" + grade + ": " + str(lovecraftChoices[lovecraft][grade])
        grandTotal+=lovecraftChoices[lovecraft][grade]
    lovecraftString = lovecraftString + "\n   -Total: " + str(grandTotal) + "/" + str(responseNum-1) + " (" + str(round(grandTotal/(responseNum-1)*100,2)) + "%)"
    resultsFile.write(lovecraftString + "\n")
resultsFile.close()

resultsFile = open("Kaiju Lovecraftian Horror Intersection Results.txt", "w", encoding='utf8')
for lovecraft in kaijuLovecraft.keys():
    kaijuString = lovecraft + ":"
    grandTotal = 0
    for kaiju in kaijuLovecraft[lovecraft].keys():
        grandTotal+=kaijuLovecraft[lovecraft][kaiju]

    for kaiju in kaijuLovecraft[lovecraft].keys():
        kaijuString = kaijuString + "\n   -" + kaiju + ": " + str(kaijuLovecraft[lovecraft][kaiju]) + " (" + str(round(kaijuLovecraft[lovecraft][kaiju]/(grandTotal)*100,2)) + "%)"
    kaijuString = kaijuString + "\n   -Total: " + str(grandTotal) + "/" + str(responseNum-1)
    resultsFile.write(kaijuString + "\n")
resultsFile.close()

for gradeLevel in grades:
    resultsFile = open(gradeLevel + "\\Kaiju Results.txt", "w", encoding='utf8')
    for kaiju in kaijuChoices.keys():
        grandTotal = 0
        for grade in kaijuChoices[kaiju].keys():
            grandTotal+=kaijuChoices[kaiju][grade]
        kaijuString = kaiju + ": " + str(kaijuChoices[kaiju][gradeLevel]) + ": " + str(grandTotal) + "/" + str(responseNum-1) + " (" + str(round(grandTotal/(responseNum-1)*100,2)) + "%)"
        resultsFile.write(kaijuString + "\n")
    resultsFile.close()


wm1 = wm.create_sheet("Kaiju")
x = 2
gradeDict = {}
b = 1
for grade in grades:
    d1 = wm1.cell(row=x, column=1,value=grade)
    gradeDict[grade] = b
    b+=1
    x+=1
x = 2
for kaiju in kaijuChoices.keys():
    d1 = wm1.cell(row=1, column=x,value=kaiju)
    for grade in kaijuChoices[kaiju].keys():
        print(grade + ": " + str(gradeDict[grade]) + "/" + str(x))
        d2 = wm1.cell(row=gradeDict[grade]+1,column=x,value=kaijuChoices[kaiju][grade])
    x+=1


wm2 = wm.create_sheet("Lovecraftian Horrors")
x = 2
gradeDict = {}
b = 1
for grade in grades:
    d1 = wm2.cell(row=x, column=1,value=grade)
    gradeDict[grade] = b
    b+=1
    x+=1
x = 2
for lovecraftianHorror in lovecraftChoices.keys():
    d1 = wm2.cell(row=1, column=x,value=lovecraftianHorror)
    for grade in lovecraftChoices[lovecraftianHorror].keys():
        print(grade + ": " + str(gradeDict[grade]) + "/" + str(x))
        d2 = wm2.cell(row=gradeDict[grade]+1,column=x,value=lovecraftChoices[lovecraftianHorror][grade])
    x+=1

wm.save("FinishedData.xlsx")
print("Completed!")