import discord
#This imports Discord. Named thing.py because my old bots had their main files in thing.js, and I'm sentimental. 
import xlrd
import xlwt
import pytz
from xlwt import Workbook
from xlutils.copy import copy 
from xlrd import open_workbook
from datetime import date
from datetime import datetime
from discord.utils import get
from openpyxl import load_workbook
import time
import matplotlib.pyplot as plt
from PIL import Image
import numpy as np

wb = Workbook()
shutDown = True
est = pytz.timezone('US/Eastern')
utc = pytz.utc
fmt = '%Y-%m-%d %H:%M:%S %Z%z'
optLevels = [-1, 0, 1, 2]
import discord
#This imports Discord. Named thing.py because my old bots had their main files in thing.js, and I'm sentimental. 
import xlrd
import xlwt
import pytz
from xlwt import Workbook
from xlutils.copy import copy 
from xlrd import open_workbook
from datetime import date
from datetime import datetime
from discord.utils import get
from openpyxl import load_workbook
import time
import matplotlib.pyplot as plt
from PIL import Image
import numpy as np

wb = Workbook()
shutDown = True
est = pytz.timezone('US/Eastern')
utc = pytz.utc
fmt = '%Y-%m-%d %H:%M:%S %Z%z'
optLevels = [-1, 0, 1, 2]

client = discord.Client()
def error():
    embed = discord.Embed(title="Error", description="Sorry, something went wrong.", color=0xFF9900)
    embed.set_footer(text="Created by The Invisible Man", icon_url="https://cdn.discordapp.com/avatars/366709133195476992/01cb7c2c7f2007d8b060e084ea4eb6fd.png?size=512")
    return embed
def errorLOCATE():
    embed = discord.Embed(title="Error", description="Sorry, we could not locate the requested user.", color=0xFF9900)
    embed.set_footer(text="Created by The Invisible Man", icon_url="https://cdn.discordapp.com/avatars/366709133195476992/01cb7c2c7f2007d8b060e084ea4eb6fd.png?size=512")
    return embed
def findUser(userID):
    return client.get_user(int(userID))
def getUserOptLevel(userID):
    optSheets = open("optSheets.txt", "r")
    optString = optSheets.read()
    optList = optString.split('\n')
    counter = -1
    for opt in optList:
        optFile = open(opt + ".txt", "r")
        optString = optFile.read()
        optArray = optString.split("\n")
        if str(userID) in optArray:
            print(str(userID) + "'s opt level is " + str(counter))
            return counter
        else:
            counter+=1
    return -5
async def on_ready():
    print('We have logged in as {0.user}'.format(client))
    print("\n")
@client.event
async def on_message(message):
    if message.author == client.user:
        return
    if message.content.startswith("&optHelp"):
        print("OptHelp request made by " + message.author.name + " at " + str(message.created_at) + " on guild " + message.guild.name)
        userOptLevel = getUserOptLevel(message.author.id)
        userOpt = "Your current Opt Level is " + str(userOptLevel) + "."
        embed = discord.Embed(title="OptHelp", description=userOpt, color=0xFF9900)
        embed.add_field(name="&optLevel -1", value="No information is collected at any point on the user. Cannot use commands.", inline=False)
        embed.add_field(name="&optLevel 0", value="Anonymus Data is collected on the user. Can use anonymus commands.", inline=False)
        embed.add_field(name="&optLevel 1", value="Data is collected in spreadsheets, but only accesible by the user. Can use all commands on self, but not on other users.", inline=False)
        embed.add_field(name="&optLevel 2", value="Data is collected; accessible by anyone. Can use all commands.", inline=False)
        embed.set_footer(text="Created by The Invisible Man", icon_url="https://cdn.discordapp.com/avatars/366709133195476992/01cb7c2c7f2007d8b060e084ea4eb6fd.png?size=512")
        await message.channel.send(embed=embed)
    #Get help relating to opt levels.
client.run('NjYyNzg4MTk1NzM3NDAzNDQy.Xg_Dqg.RvP5k1D4dWeg0tqomlXiaHz7QQg')