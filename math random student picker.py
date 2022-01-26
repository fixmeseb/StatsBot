import openpyxl

from openpyxl import load_workbook
from openpyxl import Workbook


numberOfStudents = 30
randomDigitTable = open("randomDigitTable.txt", "r").read().split("\n")
randomDigits = ""
for randomLine in randomDigitTable:
    for digitSet in randomLine.split("   "):
        randomDigits+=digitSet

wStudents = load_workbook(filename = 'US Student list.xlsx')
sheetStudents = wStudents["StudentWorkList (79)"]

emails = open("Emails.txt", "w")

grades = {
    "9th Grade" : 0,
    "10th Grade": 0,
    "11th Grade": 0,
    "12th Grade": 0
}

#9th Grade: 355-477
#10th Grade: 2-124
#11th Grade: 125-240
#12th Grade: 241-354

gradeLength = {
    1: 122,
    2: 122,
    3: 115, 
    4: 113
}


#Number: 
#9th Grade: 122
#10th Grade: 122
#11th Grade: 115
#12th Grade: 113


#122,244,366,488,610,732,854,976

#122*8=976
#122*8=976
#115*8=920
#113*8=904

position = 0

gradesString = {}

startingPosition = {
    1: 355,
    2: 2,
    3: 125,
    4: 241
}
numToWords = {
    1 : "9th Grade",
    2: "10th Grade",
    3: "11th Grade",
    4: "12th Grade"
}

randomNumberNum = 0
for number in range(1,5):
    emailList = ""
    #133-55
    gradeArray = []
    while len(gradeArray) != 8:
        #print("Random Number: " + randomDigits[position:position+3:])
        randomNumberPre = int(randomDigits[position:position+3:])
        if randomNumberPre != 0 or randomNumberPre < (gradeLength[number]*8):
            if randomNumberPre > gradeLength[number]:
                while randomNumberPre > gradeLength[number]:
                    randomNumberPre-=gradeLength[number]
            
            randomNumber = randomNumberPre + startingPosition[number]
            print(str(randomNumberPre) + " + " + str(startingPosition[number]) + " = " + str(randomNumber))

            studentName = str(sheetStudents.cell(row=randomNumber, column=2).value) + " " + str(sheetStudents.cell(row=randomNumber, column=1).value) + ", (" + str(sheetStudents.cell(row=randomNumber, column=3).value) + "): " + str(sheetStudents.cell(row=randomNumber, column=4).value)
            #print(studentName)
            grades[str(sheetStudents.cell(row=randomNumber, column=3).value)]+=1
            gradeArray.append((str(sheetStudents.cell(row=randomNumber, column=4).value) + "; "))
            randomNumberNum+=1
        else:
            "Invalid Number"
        position+=3
        #print(str(len(gradeArray)))
    gradesString[numToWords[number]] = gradeArray


for grade in grades.keys():
    print(grade + ": " + str(grades[grade]))

print("\n\n")
for grade in gradesString.keys():
    i = 1
    for person in gradesString[grade]:
        print(grade + " (" + str(i) + "): " + person)
        i+=1
        emailList +=person

emails.write(emailList[0:len(emailList)-2])
