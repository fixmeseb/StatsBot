import discord
#Looks like this searches through the #me-me-s channel on the CA Discord and writes down the number of stars per meme. Updated 12.12.2022. 

from datetime import date
from datetime import datetime
from discord.utils import get
from openpyxl import load_workbook
from openpyxl import Workbook
from pyparsing import col

import pytz

est = pytz.timezone('US/Eastern')
utc = pytz.utc

intents = discord.Intents.all()
client = discord.Client(intents=intents)
secretSecret = False

@client.event
async def on_ready():
    
    CADiscord = client.get_guild(523962430179770369)
    memes = client.get_channel(523966359877582850)
    setUp = await client.get_channel(785919963890319410).get_partial_message(785920010890903622).fetch()
    publish = setUp.created_at
    
    stars = "⭐"


    validCheckers = []
    validityMessage = await CADiscord.get_channel(524453945524092929).fetch_message(904889501720121366)
    for reaction in validityMessage.reactions:
        async for reactor in reaction.users(limit=None,after=None):
            if not (reactor.id in validCheckers):
                try:
                    bing = CADiscord.get_member(reactor.id).name
                    validCheckers.append(reactor.id)
                except:
                    bing = "incorrect"
    BotRole = CADiscord.get_role(560094358033006629)
    for member in CADiscord.members:
        if BotRole in member.roles and not(member.id in validCheckers):
            validCheckers.append(member.id)
        if secretSecret == True and not(member.id in validCheckers):
            validCheckers.append(member.id)
    
    reactorsList = ""
    for valid in validCheckers:
        reactorsList+=CADiscord.get_member(valid).name
        reactorsList+=", "
    reactorsList = reactorsList[0:len(reactorsList)-2:]
    print(reactorsList)

    memeTotalQuant = {}
    messageNumbers = {}
    for checker in validCheckers:
        memeTotalQuant[checker] = 0
    memeTotalQuant["Total"] = 0
    
    starQuant = {}
    for checker in validCheckers:
        starQuant[checker] = 0
    starQuant["Total"] = 0
    
    totalMessagesSorted = 0
    nonmemes = 0
    async for message in memes.history(limit=None,after=publish):
        #print(message.content)
        if len(message.attachments) >= 1:
            if message.author.id in validCheckers:
                memeTotalQuant[message.author.id]+=1
                #print("Added 1 message to " + str(client.get_user(message.author.id)) + " (" + str(memeTotalQuant[message.author.id]) + ")")
            memeTotalQuant["Total"]+=1
            if len(message.reactions) >=1:
                for reaction in message.reactions:
                    #print(str(message.reactions))
                    if reaction.emoji == stars and message.author.id in validCheckers:
                        starQuant[message.author.id]+=reaction.count
                        #print("Added " + str(reaction.count) + " stars to " + str(client.get_user(message.author.id)) + " (" + str(memeTotalQuant[message.author.id]) + ")")
                        starQuant["Total"]+=reaction.count
        else:
            nonmemes+=1
        
        totalMessagesSorted+=1
        if totalMessagesSorted % 500 == 0:
            print(str(totalMessagesSorted) + " messages sorted.")

    print("Total memes: " + str(memeTotalQuant["Total"]))
    print("Non-memes: " + str(nonmemes))
    print("Total Stars: " + str(starQuant["Total"]))
    wb = Workbook()
    ws2 = wb.create_sheet("Star Counts")
    ws1 = wb.create_sheet("Message Counts")
    print("Totals: ")
    rowNum = 1
    rowNumMessages = 1
    rowNumStars = 1
    for author in memeTotalQuant.keys():
        print(author)
        try:
            ws1.cell(row=rowNumMessages,column=1,value=str(client.get_user(author)))
            ws2.cell(row=rowNumStars,column=1,value=str(client.get_user(author)))
        except:
            print("Author = " + str(author))
            ws1.cell(row=rowNumMessages,column=1,value=author)
            ws2.cell(row=rowNumStars,column=1,value=author)
        ws1.cell(row=rowNumMessages,column=2,value=memeTotalQuant[author])
        rowNumMessages+=1
        ws2.cell(row=rowNumStars,column=2,value=starQuant[author])
        rowNumStars+=1
    wb.save("Stars Per Meme.xlsx")


    print("Completed!")

client.run('NjYyNzg4MTk1NzM3NDAzNDQy.Gw3sKx.6XM0TZPT6SP5i4vyJ4KoYZlfp_eeOhgGA8icLc')