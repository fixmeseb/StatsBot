import win32com.client
#other libraries to be used in this script
import openpyxl
from openpyxl import Workbook


yearsSenders = {}
totalNum = 0
senders = {}

aliases = {
    "tamipolge@yahoo.com": "Tami Polge",
    "tami.polge@gmail.com": "Tami Polge",
    "m@mail4.veracross.com": "Veracross",
    "m@mail2.veracross.com": "Veracross",
    "m@mail1.veracross.com": "Veracross",
    "m@mail3.veracross.com": "Veracross",
    "steve.polge@epicgames.com": "Steven Polge",
    "steve.polge@gmail.com": "Steven Polge"
}

blanks = []

outlook = win32com.client.Dispatch('outlook.application')
mapi = outlook.GetNamespace("MAPI")
errors = []

for account in mapi.Accounts:
	print(account.DeliveryStore.DisplayName)
print("\n")
inbox = mapi.GetDefaultFolder(6)

speedRun = False
if speedRun == False:
    for folder in list(mapi.GetDefaultFolder(6).Folders):
        print(str(folder))
        messages = folder.Items

        try:
            for subfolder in list(folder.folders):
                print(str(subfolder))
                try:
                    for subsubfolder in list(subfolder.folders):
                        print(str(subsubfolder))
                        messages = subsubfolder.Items

                        i = 0
                        #try:
                        if True == True:
                            for message in list(messages):
                                try:
                                    #print(str(message.subject) + message.body)
                                    if str(message.subject) != None:
                                        #print(str(message.SenderEmailAddress) + ": " + str(message.subject))
                                        name = "Miracle Mike and the Miraculous Miracle Band"
                                        try:
                                            name = str(message.SenderEmailAddress)
                                            if message.SenderEmailAddress[0:16:] == "/O=EXCHANGELABS/" or message.SenderEmailAddress[0:16:] == "/o=ExchangeLabs/":
                                                name = str(message.sender)
                                        except:
                                            name = str(message.Organizer)
                                        try:
                                            name = aliases[name]
                                        except:
                                            baa = False
                                        
                                        #if message.SenderEmailAddress[0:16:] == "/O=EXCHANGELABS/"
                                        if name != "":
                                            try:
                                                senders[name]+=1
                                            except:
                                                senders[name] = 1
                                            try:
                                                yearsSenders[name][message.ReceivedTime.year]+=1
                                            except:
                                                yearsSenders[name] = {
                                                    2015: 0,
                                                    2016: 0,
                                                    2017: 0,
                                                    2018: 0,
                                                    2019: 0,
                                                    2020: 0,
                                                    2021: 0,
                                                    2022: 0
                                                }
                                                yearsSenders[name][message.ReceivedTime.year]+=1
                                            
                                        if name == "":
                                            blanks.append(message.subject)
                                        i+=1
                                        totalNum+=1
                                        if totalNum % 500 == 0:
                                            print(str(i) + "/" + str(totalNum) + " messages sorted!")
                                            #print(str(name) + ": " + str(message.subject))
                                except:
                                    print(str(folder) + "\\" + str(subfolder) + "\\" + str(subsubfolder) + " (subsubfolder): Error in message: " + str(message.subject))
                                    errors.append(str(folder) + "\\" + str(subfolder) + "\\" + str(subsubfolder) + " (subsubfolder):\t" + str(message.subject))
                            print("Funished Successfully!")
                except:
                    print("Subsubfolder error: " + str(folder) + " has no subsubfolders")
                i = 0
                messages = subfolder.Items

                #try:
                if True == True:
                    for message in list(messages):
                        try:
                            #print(str(message.subject) + message.body)
                            name = "Miracle Mike and the Miraculous Miracle Band"
                            try:
                                name = str(message.SenderEmailAddress)
                                if message.SenderEmailAddress[0:16:] == "/O=EXCHANGELABS/" or message.SenderEmailAddress[0:16:] == "/o=ExchangeLabs/":
                                    name = str(message.sender)
                            except:
                                name = str(message.Organizer)
                                #if message.SenderEmailAddress[0:16:] == "/O=EXCHANGELABS/"
                            try:
                                name = aliases[name]
                            except:
                                baa = False
                            if name != "":
                                try:
                                    senders[name]+=1
                                except:
                                    senders[name] = 1
                                try:
                                    yearsSenders[name][message.ReceivedTime.year]+=1
                                except:
                                    yearsSenders[name] = {
                                        2015: 0,
                                        2016: 0,
                                        2017: 0,
                                        2018: 0,
                                        2019: 0,
                                        2020: 0,
                                        2021: 0,
                                        2022: 0
                                    }
                                    yearsSenders[name][message.ReceivedTime.year]+=1

                            if name == "":
                                blanks.append(message.subject)
                            i+=1
                            totalNum+=1
                            if totalNum % 500 == 0:
                                print(str(i) + "/" + str(totalNum) + " messages sorted!")
                                #print(str(name) + ": " + str(message.subject))
                        except:
                            
                            print(str(folder) + "\\" + str(subfolder) + ": (subfolder) Error in message: " + str(message.subject))
                            errors.append(str(folder) + "\\" + str(subfolder) + ": (subfolder): \t" + str(message.subject))

                    print("Funished Successfully!")
        except:
            print("Subfolder error: " + str(folder) + " has no subfolders")
        i = 0
        #try:
        if True == True:
            for message in list(messages):
                try:
                    #print(str(message.subject) + message.body)
                    if str(message.subject) != None:
                        #print(str(message.SenderEmailAddress) + ": " + str(message.subject))
                        name = "Miracle Mike and the Miraculous Miracle Band"
                        try:
                            name = str(message.SenderEmailAddress)
                            if message.SenderEmailAddress[0:16:] == "/O=EXCHANGELABS/" or message.SenderEmailAddress[0:16:] == "/o=ExchangeLabs/":
                                name = str(message.sender)
                        except:
                            name = str(message.Organizer)
                            #if message.SenderEmailAddress[0:16:] == "/O=EXCHANGELABS/"
                        try:
                            name = aliases[name]
                        except:
                            baa = False
                        if name != "":
                            try:
                                senders[name]+=1
                            except:
                                senders[name] = 1
                            try:
                                yearsSenders[name][message.ReceivedTime.year]+=1
                            except:
                                yearsSenders[name] = {
                                    2015: 0,
                                    2016: 0,
                                    2017: 0,
                                    2018: 0,
                                    2019: 0,
                                    2020: 0,
                                    2021: 0,
                                    2022: 0
                                }
                                yearsSenders[name][message.ReceivedTime.year]+=1
                        if name == "":
                            blanks.append(message.subject)
                        i+=1
                        totalNum+=1
                        if totalNum % 500 == 0:
                            print(str(i) + "/" + str(totalNum) + " messages sorted!")
                            #print(str(name) + ": " + str(message.subject))

                except:

                    print(str(folder) + ": (folder) Error in message: " + str(message.subject))
                    errors.append(str(folder) + ": (folder): \t" + str(message.subject))

            print("Funished Successfully!")
    #inbox = mapi.GetDefaultFolder(6).Folders["your_sub_folder"]

messages = inbox.Items

i = 0
#try:
if True == True:
    for message in list(messages):
        try:
            #print(str(message.subject) + message.body)
            if str(message.subject) != None:
                #print(str(message.SenderEmailAddress) + ": " + str(message.subject))
                name = "Miracle Mike and the Miraculous Miracle Band"
                try:
                    name = str(message.SenderEmailAddress)
                    if message.SenderEmailAddress[0:16:] == "/O=EXCHANGELABS/" or message.SenderEmailAddress[0:16:] == "/o=ExchangeLabs/":
                        name = str(message.sender)
                except:
                    name = str(message.Organizer)
                #print(str(message.SenderEmailAddress) + "/" + str(name) + ": " + str(message.subject))
                    #if message.SenderEmailAddress[0:16:] == "/O=EXCHANGELABS/"
                try:
                    name = aliases[name]
                except:
                    baa = False
                if name != "":
                    try:
                        senders[name]+=1
                    except:
                        senders[name] = 1
                    try:
                        yearsSenders[name][message.ReceivedTime.year]+=1
                    except:
                        yearsSenders[name] = {
                            2015: 0,
                            2016: 0,
                            2017: 0,
                            2018: 0,
                            2019: 0,
                            2020: 0,
                            2021: 0,
                            2022: 0
                        }
                        yearsSenders[name][message.ReceivedTime.year]+=1

                                            
                if name == "":
                    blanks.append(message.subject)
                #print(str(message.ReceivedTime.year))
                i+=1
                totalNum+=1
                if totalNum % 500 == 0:
                    print(str(i) + "/" + str(totalNum) + " messages sorted!")
                    #print(str(name) + ": " + str(message.subject))
            if str(message.subject) == "CATF - today's practice plans":
                print(str(i) + ": " + str(message.SenderEmailAddress) + ": " + str(message.subject))

        except:
            print("Inbox: Error in message!" + str(message.subject))
            errors.append("Inbox: " + str(message.subject))

    print("Funished Successfully!")

wb = Workbook()
ws = wb.active

wb2 = Workbook()
ws2 = wb2.active

values = {}
for person in senders.keys():
    try:
        values[senders[person]].append(person)
    except:
        values[senders[person]] = [person]

valuesAway = list(values.keys())
valuesAway.sort(reverse=True)
#people = list(senders.keys())
#people.sort()
#print(str(people))

print(str(yearsSenders))
print(str(yearsSenders["Conrad Hall"]))

rowNum = 1
colNum = 1

for value in valuesAway: 
    checkSort = values[value]
    checkSort.sort()
    for valueMan in checkSort:
        ws.cell(column=1,row=rowNum,value=str(valueMan))
        ws.cell(column=2,row=rowNum,value=str(value))
        colNum = 2
        for year in yearsSenders[valueMan].keys():
            ws2.cell(column=1,row=rowNum+1,value=str(valueMan))
            ws2.cell(column=colNum,row=rowNum+1,value=str(yearsSenders[valueMan][year]))
            colNum+=1
        rowNum+=1

wb.save('Email Count.xlsx')
wb2.save("Email Count Years.xlsx")
#except:
	#print("error when processing emails messages")
print(str(totalNum) + " messages!")

print("Completed!")