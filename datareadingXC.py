import os
import datetime
import openpyxl

from openpyxl import load_workbook
from openpyxl import Workbook

def convertToSec(time):
    minutes = int(time.split(":")[0])
    seconds = int(time.split(":")[1].split(".")[0])
    nano = int(time.split(":")[1].split(".")[1])

    timeFinal = float(str(minutes*60 + seconds) + "." + str(nano))
    return(timeFinal)


class Person:
    def __init__(self):
        self.time = ""
        self.school = ""
        self.lastName = ""
        self.firstName = ""
        self.grade = 0
        """
        :param string time: the time that the runner ran.
        """

class Race:
    def __init__(self):
        self.people = []
        self.courseName = ""
        self.date = datetime.date.today()
        self.name = ""

class Racer:
    def __init__(self):
        self.firstName = ""
        self.lastName = ""
        self.school = ""
        self.times = {}

foo = 1
people = {}

courses = ["Hagan Stone", "Ivey Redmond", "McAlpine Park (Larry McAfee)", "Wakemed"]
courseData = {}
debug = False
if debug == False:
    try:
        for raceCourse in courses:
            basepath = 'C:\\Users\\Sebastian_Polge\\OneDrive - Cary Academy\\Desktop\\XC Data\\' + raceCourse
            for entry in os.listdir(basepath):
                if os.path.isfile(os.path.join(basepath, entry)):
                    dataFile = open(basepath + "\\" + entry, "r", encoding='utf8')
                    dataFull = dataFile.read()
                    currentRace = Race()
                    currentRace.people = []
                    if dataFull.split("\n")[5] == "================================================================================" and dataFull.split("\n")[2].strip() == "Results":
                        print("Type 1 Exists for " + entry)
                        dataArray = dataFull.split("================================================================================")
                        data = []
                        for dataSegmentNum in range(len(dataArray)):
                            if dataSegmentNum != 0 and dataSegmentNum % 2 == 0:
                                info = dataArray[dataSegmentNum]
                                infoArray = info.split("\n")[0:len(info)-2:]
                                for personData in infoArray:
                                    try:
                                        x = int(personData[0:3:].strip())
                                        data.append(personData)
                                    except:
                                        boolean = False
                        for personString in data:
                            newPerson = Person()
                            names = personString[4:32:].strip().split(", ")
                            
                            newPerson.lastName = names[0]
                            newPerson.firstName = names[1]
                            newPerson.school = personString[33:45:]
                            newPerson.time = personString[64:72:]
                            currentRace.people.append(newPerson)
                            if newPerson.firstName == "Sebastian" and newPerson.lastName == "Polge":
                                print("added " + newPerson.firstName + " " + newPerson.lastName)
                        currentRace.courseName = dataFull.split("\n")[1].strip()
                        raceInfo = dataFull.split("\n")[0].split("-")
                        currentRace.name = raceInfo[0].strip()
                        raceDateString = raceInfo[1].strip().split("/")
                        currentRace.date = datetime.date(int(raceDateString[2]), int(raceDateString[0]), int(raceDateString[1]))
                        courseData[currentRace.name] = currentRace
                    else:
                        if dataFull.split("\n")[6] == "-----------------------------------------------------------------":
                            print("Type 2 Exists for " + entry)
                            dataArray = dataFull.split("-----------------------------------------------------------------")
                            data = []
                            for info in dataArray[1].split("\n"):
                                try:
                                    x = int(info[0:3:].strip())
                                    #print(str(x))
                                    data.append(info)
                                except:
                                    boolean = False
                            #print(data)
                            for line in data:
                                if line != "":
                                    for personString in data:
                                        newPerson = Person()
                                        #print(personString)
                                        names = personString[4:26:].strip().split(" ")
                                        #print(names)
                                        newPerson.lastName = names[1]
                                        newPerson.firstName = names[0]
                                        try:
                                            grade = personString[26:29:].strip()
                                            newPerson.grade = int(grade)
                                        except:
                                            boolean = False
                                            #print(newPerson.firstName + " " + newPerson.lastName + " is missing a grade.")                            
                                        newPerson.school = personString[29:50:].strip()
                                        newPerson.time = personString[50:60:].strip()
                                        #print(newPerson.firstName + " " + newPerson.lastName + ": " + newPerson.school + ", " + newPerson.time + " in " + str(newPerson.grade) + "th grade.")
                                        currentRace.people.append(newPerson)
                                        if newPerson.firstName == "Sebastian" and newPerson.lastName == "Polge":
                                            print("added " + newPerson.firstName + " " + newPerson.lastName)
                            currentRace.courseName = dataFull.split("\n")[1].strip()
                            raceInfo = dataFull.split("\n")[0].split("-")
                            currentRace.name = raceInfo[0].strip()
                            raceDateString = dataFull.split("\n")[0].split("/")
                            #print("Month: " + str(raceDateString[0][len(raceDateString[0])-2:len(raceDateString[0]):].strip()))
                            #print("Day: " + str(raceDateString[1]))
                            #print("Year: " + str(raceDateString[2][0:5:]))
                            #print(str(raceDateString[0][len(raceDateString[0])-2:len(raceDateString[0]):].strip()) + "/" + str(raceDateString[1]) + "/" + str(raceDateString[2][0:5:]))
                            currentRace.date = datetime.date(int(raceDateString[2][0:5:]), int(raceDateString[0][len(raceDateString[0])-2:len(raceDateString[0]):].strip()), int(raceDateString[1]))
                            courseData[currentRace.name] = currentRace
                            #print(currentRace.name + ": hosted at " + currentRace.courseName + " on " + str(currentRace.date))
                        else:
                            if False == True:
                                print("Type 5 Exists for " + entry)
                            else:
                                if dataFull.split("\n")[2 ] == "" and dataFull.split("\n")[4].strip() == "===================================================================================================":
                                    print("Type 4 Exists for " + entry)
                                    dataArray = dataFull.split("================================================================================")
                                    data = []
                                    nextGood = False
                                    for dataChunk in dataArray:
                                        if nextGood == True:
                                            nextGood = False
                                            dataSplits = dataChunk.split("\n")
                                            
                                            for dataLine in dataSplits:
                                                try:
                                                    x = int(dataLine[0:5:].strip())
                                                    data.append(dataLine)
                                                except:
                                                    boolean = False
                                        if dataChunk.strip().startswith("Pl Athlete") or dataChunk.strip()[0:10] == "Pl Athlete":
                                            nextGood = True
                                        
                                    for personString in data:
                                        newPerson = Person()
                                        names = personString[4:29:].strip().split(", ")
                                        
                                        newPerson.lastName = names[0][0:1:].upper() + names[0][1:len(names[0])].lower()
                                        try:
                                            newPerson.firstName = names[1]
                                        except:
                                            boolean = False
                                        newPerson.school = personString[37:72:].strip()
                                        newPerson.time = personString[72:79:].strip()
                                        grades = {
                                            "5": 5,
                                            "6": 6,
                                            "7": 7,
                                            "8": 8,
                                            "9": 9,
                                            "10": 10,
                                            "11": 11,
                                            "12": 12,
                                            "13": 13,
                                            "15": 15,
                                            "17": 17,
                                            "FR": 9,
                                            "SO": 10,
                                            "JR": 11,
                                            "SR": 12
                                        }
                                        gradeRude = personString[29:32].strip()
                                        if gradeRude != "":
                                            newPerson.grade = grades[gradeRude]

                                        currentRace.people.append(newPerson)
                                        if newPerson.firstName == "Sebastian" and newPerson.lastName == "Polge":
                                            print("added " + newPerson.firstName + " " + newPerson.lastName)
                                    currentRace.courseName = dataFull.split("\n")[1].strip().split(" - ")[0]
                                    raceInfo = dataFull.split("\n")
                                    currentRace.name = dataFull.split("\n")[0].strip()
                                    raceDateString = dataFull.split("\n")[1].strip().split(" - ")[1].strip().split("/")
                                    currentRace.date = datetime.date(int(raceDateString[2]), int(raceDateString[0]), int(raceDateString[1]))
                                    courseData[currentRace.name] = currentRace
                                else:
                                    if dataFull.split("\n")[2 ] == "" and dataFull.split("\n")[3].strip() == "================================================================================":
                                        print("Type 3 Exists for " + entry)
                                        dataArray = dataFull.split("================================================================================")
                                        data = []
                                        nextGood = False
                                        for dataChunk in dataArray:
                                            if nextGood == True:
                                                nextGood = False
                                                dataSplits = dataChunk.split("\n")
                                                
                                                for dataLine in dataSplits:
                                                    try:
                                                        x = int(dataLine[0:5:].strip())
                                                        data.append(dataLine)
                                                    except:
                                                        boolean = False
                                            if dataChunk.strip().startswith("Pl Athlete") or dataChunk.strip()[0:10] == "Pl Athlete":
                                                nextGood = True
                                            
                                        for personString in data:
                                            newPerson = Person()
                                            names = personString[4:29:].strip().split(", ")
                                            
                                            newPerson.lastName = names[0][0:1:].upper() + names[0][1:len(names[0])].lower()
                                            try:
                                                newPerson.firstName = names[1]
                                            except:
                                                boolean = False
                                            newPerson.school = personString[37:66:].strip()
                                            newPerson.time = personString[72:79:].strip()
                                            grades = {
                                                "5": 5,
                                                "6": 6,
                                                "7": 7,
                                                "8": 8,
                                                "9": 9,
                                                "10": 10,
                                                "11": 11,
                                                "12": 12,
                                                "13": 13,
                                                "15": 15,
                                                "17": 17,
                                                "FR": 9,
                                                "SO": 10,
                                                "JR": 11,
                                                "SR": 12
                                            }
                                            gradeRude = personString[29:32].strip()
                                            if gradeRude != "":
                                                newPerson.grade = grades[gradeRude]

                                            currentRace.people.append(newPerson)
                                            if newPerson.firstName == "Sebastian" and newPerson.lastName == "Polge":
                                                print("added " + newPerson.firstName + " " + newPerson.lastName)
                                        currentRace.courseName = dataFull.split("\n")[1].strip().split(" - ")[0]
                                        raceInfo = dataFull.split("\n")
                                        currentRace.name = dataFull.split("\n")[0].strip()
                                        raceDateString = dataFull.split("\n")[1].strip().split(" - ")[1].strip().split("/")
                                        currentRace.date = datetime.date(int(raceDateString[2]), int(raceDateString[0]), int(raceDateString[1]))
                                        courseData[currentRace.name] = currentRace
                                    else:
                                        if dataFull.split("\n")[3].strip() == "Results" and dataFull.split("\n")[6].strip() == "=======================================================================":
                                            print("Type 6 Exists for " + entry)
                                            dataArray = dataFull.split("=======================================================================")
                                            data = []
                                            nextGood = False
                                            for dataChunk in dataArray:
                                                if nextGood == True:
                                                    nextGood = False
                                                    dataSplits = dataChunk.split("\n")
                                                    
                                                    for dataLine in dataSplits:
                                                        try:
                                                            x = int(dataLine[0:4:].strip())
                                                            data.append(dataLine)
                                                        except:
                                                            boolean = False
                                                if dataChunk.strip().startswith("Name") or dataChunk.strip()[0:4] == "Name":
                                                    nextGood = True
                                                
                                            output = open("Debug.txt", "w")
                                            for personString in data:
                                                newPerson = Person()
                                                names = personString[4:30:].strip().split(" ")
                                                
                                                newPerson.firstName = names[0]
                                                try:
                                                    newPerson.lastName = names[1]
                                                except:
                                                    boolean = False
                                                newPerson.school = personString[33:56:].strip()
                                                newPerson.time = personString[56:63:].strip()
                                                
                                                gradeRude = personString[30:33].strip()
                                                if gradeRude != "":
                                                    newPerson.grade = grades[gradeRude]
                                                currentRace.people.append(newPerson)
                                                if newPerson.firstName == "Sebastian" and newPerson.lastName == "Polge":
                                                    print("added " + newPerson.firstName + " " + newPerson.lastName)
                                                output.write(str(len(currentRace.people)) + ". " + newPerson.firstName + " " + newPerson.lastName + "\n")
                                                
                                            raceInfo = dataFull.split("\n")
                                            currentRace.name = raceInfo[0].strip().split(" - ")[0]
                                            currentRace.courseName = raceInfo[2].strip()
                                            raceDateString = raceInfo[0].strip().split(" - ")[1].strip().split(" to ")[0].strip().split("/")
                                            currentRace.date = datetime.date(int(raceDateString[2]), int(raceDateString[0]), int(raceDateString[1]))
                                            courseData[currentRace.name] = currentRace
                                            output2 = open("Debug2.txt", "w")
                                            for person in currentRace.people:
                                                output2.write(person.firstName + " " + person.lastName + "\n")
                                        else:
                                            print("Unsorted Entry: " + entry)
                checkFile2 = open("OutputDebug" + str(foo) + ".txt", "w")

                for person in courseData["37th Annual Greensboro XC Invitational 2021"].people:
                    checkFile2.write(person.firstName + " " + person.lastName + "\n")
                print("created OutputDebug" + str(foo) + ".txt")
                checkFile2.close()
                foo+=1
    except:
        for raceCourse in courses:
            basepath = 'C:\\Users\\Admin\\OneDrive - Cary Academy\\Desktop\\XC Data\\' + raceCourse
            for entry in os.listdir(basepath):
                if os.path.isfile(os.path.join(basepath, entry)):
                    dataFile = open(basepath + "\\" + entry, "r", encoding='utf8')
                    dataFull = dataFile.read()
                    currentRace = Race()
                    currentRace.people = []
                    if dataFull.split("\n")[5] == "================================================================================" and dataFull.split("\n")[2].strip() == "Results":
                        print("Type 1 Exists for " + entry)
                        dataArray = dataFull.split("================================================================================")
                        data = []
                        for dataSegmentNum in range(len(dataArray)):
                            if dataSegmentNum != 0 and dataSegmentNum % 2 == 0:
                                info = dataArray[dataSegmentNum]
                                infoArray = info.split("\n")[0:len(info)-2:]
                                for personData in infoArray:
                                    try:
                                        x = int(personData[0:3:].strip())
                                        data.append(personData)
                                    except:
                                        boolean = False
                        for personString in data:
                            newPerson = Person()
                            names = personString[4:32:].strip().split(", ")
                            
                            newPerson.lastName = names[0]
                            newPerson.firstName = names[1]
                            newPerson.school = personString[33:45:]
                            newPerson.time = personString[64:72:]
                            currentRace.people.append(newPerson)
                            if newPerson.firstName == "Sebastian" and newPerson.lastName == "Polge":
                                print("added " + newPerson.firstName + " " + newPerson.lastName)
                        currentRace.courseName = dataFull.split("\n")[1].strip()
                        raceInfo = dataFull.split("\n")[0].split("-")
                        currentRace.name = raceInfo[0].strip()
                        raceDateString = raceInfo[1].strip().split("/")
                        currentRace.date = datetime.date(int(raceDateString[2]), int(raceDateString[0]), int(raceDateString[1]))
                        courseData[currentRace.name] = currentRace
                    else:
                        if dataFull.split("\n")[6] == "-----------------------------------------------------------------":
                            print("Type 2 Exists for " + entry)
                            dataArray = dataFull.split("-----------------------------------------------------------------")
                            data = []
                            for info in dataArray[1].split("\n"):
                                try:
                                    x = int(info[0:3:].strip())
                                    #print(str(x))
                                    data.append(info)
                                except:
                                    boolean = False
                            #print(data)
                            for line in data:
                                if line != "":
                                    for personString in data:
                                        newPerson = Person()
                                        #print(personString)
                                        names = personString[4:26:].strip().split(" ")
                                        #print(names)
                                        newPerson.lastName = names[1]
                                        newPerson.firstName = names[0]
                                        try:
                                            grade = personString[26:29:].strip()
                                            newPerson.grade = int(grade)
                                        except:
                                            boolean = False
                                            #print(newPerson.firstName + " " + newPerson.lastName + " is missing a grade.")                            
                                        newPerson.school = personString[29:50:].strip()
                                        newPerson.time = personString[50:60:].strip()
                                        #print(newPerson.firstName + " " + newPerson.lastName + ": " + newPerson.school + ", " + newPerson.time + " in " + str(newPerson.grade) + "th grade.")
                                        currentRace.people.append(newPerson)
                                        if newPerson.firstName == "Sebastian" and newPerson.lastName == "Polge":
                                            print("added " + newPerson.firstName + " " + newPerson.lastName)
                            currentRace.courseName = dataFull.split("\n")[1].strip()
                            raceInfo = dataFull.split("\n")[0].split("-")
                            currentRace.name = raceInfo[0].strip()
                            raceDateString = dataFull.split("\n")[0].split("/")
                            #print("Month: " + str(raceDateString[0][len(raceDateString[0])-2:len(raceDateString[0]):].strip()))
                            #print("Day: " + str(raceDateString[1]))
                            #print("Year: " + str(raceDateString[2][0:5:]))
                            #print(str(raceDateString[0][len(raceDateString[0])-2:len(raceDateString[0]):].strip()) + "/" + str(raceDateString[1]) + "/" + str(raceDateString[2][0:5:]))
                            currentRace.date = datetime.date(int(raceDateString[2][0:5:]), int(raceDateString[0][len(raceDateString[0])-2:len(raceDateString[0]):].strip()), int(raceDateString[1]))
                            courseData[currentRace.name] = currentRace
                            #print(currentRace.name + ": hosted at " + currentRace.courseName + " on " + str(currentRace.date))
                        else:
                            if False == True:
                                print("Type 5 Exists for " + entry)
                            else:
                                if dataFull.split("\n")[2 ] == "" and dataFull.split("\n")[4].strip() == "===================================================================================================":
                                    print("Type 4 Exists for " + entry)
                                    dataArray = dataFull.split("================================================================================")
                                    data = []
                                    nextGood = False
                                    for dataChunk in dataArray:
                                        if nextGood == True:
                                            nextGood = False
                                            dataSplits = dataChunk.split("\n")
                                            
                                            for dataLine in dataSplits:
                                                try:
                                                    x = int(dataLine[0:5:].strip())
                                                    data.append(dataLine)
                                                except:
                                                    boolean = False
                                        if dataChunk.strip().startswith("Pl Athlete") or dataChunk.strip()[0:10] == "Pl Athlete":
                                            nextGood = True
                                        
                                    for personString in data:
                                        newPerson = Person()
                                        names = personString[4:29:].strip().split(", ")
                                        
                                        newPerson.lastName = names[0][0:1:].upper() + names[0][1:len(names[0])].lower()
                                        try:
                                            newPerson.firstName = names[1]
                                        except:
                                            boolean = False
                                        newPerson.school = personString[37:72:].strip()
                                        newPerson.time = personString[72:79:].strip()
                                        grades = {
                                            "5": 5,
                                            "6": 6,
                                            "7": 7,
                                            "8": 8,
                                            "9": 9,
                                            "10": 10,
                                            "11": 11,
                                            "12": 12,
                                            "13": 13,
                                            "15": 15,
                                            "17": 17,
                                            "FR": 9,
                                            "SO": 10,
                                            "JR": 11,
                                            "SR": 12
                                        }
                                        gradeRude = personString[29:32].strip()
                                        if gradeRude != "":
                                            newPerson.grade = grades[gradeRude]

                                        currentRace.people.append(newPerson)
                                        if newPerson.firstName == "Sebastian" and newPerson.lastName == "Polge":
                                            print("added " + newPerson.firstName + " " + newPerson.lastName)
                                    currentRace.courseName = dataFull.split("\n")[1].strip().split(" - ")[0]
                                    raceInfo = dataFull.split("\n")
                                    currentRace.name = dataFull.split("\n")[0].strip()
                                    raceDateString = dataFull.split("\n")[1].strip().split(" - ")[1].strip().split("/")
                                    currentRace.date = datetime.date(int(raceDateString[2]), int(raceDateString[0]), int(raceDateString[1]))
                                    courseData[currentRace.name] = currentRace
                                else:
                                    if dataFull.split("\n")[2 ] == "" and dataFull.split("\n")[3].strip() == "================================================================================":
                                        print("Type 3 Exists for " + entry)
                                        dataArray = dataFull.split("================================================================================")
                                        data = []
                                        nextGood = False
                                        for dataChunk in dataArray:
                                            if nextGood == True:
                                                nextGood = False
                                                dataSplits = dataChunk.split("\n")
                                                
                                                for dataLine in dataSplits:
                                                    try:
                                                        x = int(dataLine[0:5:].strip())
                                                        data.append(dataLine)
                                                    except:
                                                        boolean = False
                                            if dataChunk.strip().startswith("Pl Athlete") or dataChunk.strip()[0:10] == "Pl Athlete":
                                                nextGood = True
                                            
                                        for personString in data:
                                            newPerson = Person()
                                            names = personString[4:29:].strip().split(", ")
                                            
                                            newPerson.lastName = names[0][0:1:].upper() + names[0][1:len(names[0])].lower()
                                            try:
                                                newPerson.firstName = names[1]
                                            except:
                                                boolean = False
                                            newPerson.school = personString[37:66:].strip()
                                            newPerson.time = personString[72:79:].strip()
                                            grades = {
                                                "5": 5,
                                                "6": 6,
                                                "7": 7,
                                                "8": 8,
                                                "9": 9,
                                                "10": 10,
                                                "11": 11,
                                                "12": 12,
                                                "13": 13,
                                                "15": 15,
                                                "17": 17,
                                                "FR": 9,
                                                "SO": 10,
                                                "JR": 11,
                                                "SR": 12
                                            }
                                            gradeRude = personString[29:32].strip()
                                            if gradeRude != "":
                                                newPerson.grade = grades[gradeRude]

                                            currentRace.people.append(newPerson)
                                            if newPerson.firstName == "Sebastian" and newPerson.lastName == "Polge":
                                                print("added " + newPerson.firstName + " " + newPerson.lastName)
                                        currentRace.courseName = dataFull.split("\n")[1].strip().split(" - ")[0]
                                        raceInfo = dataFull.split("\n")
                                        currentRace.name = dataFull.split("\n")[0].strip()
                                        raceDateString = dataFull.split("\n")[1].strip().split(" - ")[1].strip().split("/")
                                        currentRace.date = datetime.date(int(raceDateString[2]), int(raceDateString[0]), int(raceDateString[1]))
                                        courseData[currentRace.name] = currentRace
                                    else:
                                        if dataFull.split("\n")[3].strip() == "Results" and dataFull.split("\n")[6].strip() == "=======================================================================":
                                            print("Type 6 Exists for " + entry)
                                            dataArray = dataFull.split("=======================================================================")
                                            data = []
                                            nextGood = False
                                            for dataChunk in dataArray:
                                                if nextGood == True:
                                                    nextGood = False
                                                    dataSplits = dataChunk.split("\n")
                                                    
                                                    for dataLine in dataSplits:
                                                        try:
                                                            x = int(dataLine[0:4:].strip())
                                                            data.append(dataLine)
                                                        except:
                                                            boolean = False
                                                if dataChunk.strip().startswith("Name") or dataChunk.strip()[0:4] == "Name":
                                                    nextGood = True
                                                
                                            output = open("Debug.txt", "w")
                                            for personString in data:
                                                newPerson = Person()
                                                names = personString[4:30:].strip().split(" ")
                                                
                                                newPerson.firstName = names[0]
                                                try:
                                                    newPerson.lastName = names[1]
                                                except:
                                                    boolean = False
                                                newPerson.school = personString[33:56:].strip()
                                                newPerson.time = personString[56:63:].strip()
                                                
                                                gradeRude = personString[30:33].strip()
                                                if gradeRude != "":
                                                    newPerson.grade = grades[gradeRude]
                                                currentRace.people.append(newPerson)
                                                if newPerson.firstName == "Sebastian" and newPerson.lastName == "Polge":
                                                    print("added " + newPerson.firstName + " " + newPerson.lastName)
                                                output.write(str(len(currentRace.people)) + ". " + newPerson.firstName + " " + newPerson.lastName + "\n")
                                                
                                            raceInfo = dataFull.split("\n")
                                            currentRace.name = raceInfo[0].strip().split(" - ")[0]
                                            currentRace.courseName = raceInfo[2].strip()
                                            raceDateString = raceInfo[0].strip().split(" - ")[1].strip().split(" to ")[0].strip().split("/")
                                            currentRace.date = datetime.date(int(raceDateString[2]), int(raceDateString[0]), int(raceDateString[1]))
                                            courseData[currentRace.name] = currentRace
                                            output2 = open("Debug2.txt", "w")
                                            for person in currentRace.people:
                                                output2.write(person.firstName + " " + person.lastName + "\n")
                                        else:
                                            print("Unsorted Entry: " + entry)
                checkFile2 = open("OutputDebug" + str(foo) + ".txt", "w")

                for person in courseData["37th Annual Greensboro XC Invitational 2021"].people:
                    checkFile2.write(person.firstName + " " + person.lastName + "\n")
                print("created OutputDebug" + str(foo) + ".txt")
                checkFile2.close()
                foo+=1
else:
    basepath = 'C:\\Users\\Sebastian_Polge\\OneDrive - Cary Academy\\Desktop\\XC Data\\Hagan Stone'
    for entry in os.listdir(basepath):
        if os.path.isfile(os.path.join(basepath, entry)):
            dataFile = open(basepath + "\\" + entry, "r", encoding='utf8')
            dataFull = dataFile.read()
            #print(dataFull)
            currentRace = Race()
            if dataFull.split("\n")[5] == "================================================================================" and dataFull.split("\n")[2].strip() == "Results":
                print("Type 1 Exists for " + entry)
                dataArray = dataFull.split("================================================================================")
                data = []
                for dataSegmentNum in range(len(dataArray)):
                    if dataSegmentNum != 0 and dataSegmentNum % 2 == 0:
                        info = dataArray[dataSegmentNum]
                        infoArray = info.split("\n")[0:len(info)-2:]
                        #print("info")
                        for personData in infoArray:
                            try:
                                x = int(personData[0:3:].strip())
                                #print(personData)
                                data.append(personData)
                            except:
                                boolean = False
                #print(str(data))
                for personString in data:
                    newPerson = Person()
                    names = personString[4:32:].strip().split(", ")
                    #print(personString)
                    #print(names)
                    newPerson.lastName = names[0]
                    newPerson.firstName = names[1]
                    newPerson.school = personString[33:45:].strip()
                    newPerson.time = personString[64:72:].strip()
                    print(newPerson.firstName + " " + newPerson.lastName + ": " + newPerson.school + ", " + newPerson.time)
                    currentRace.people.append(newPerson)
                    if newPerson.firstName == "Sebastian" and newPerson.lastName == "Polge":
                        print("added " + newPerson.firstName + " " + newPerson.lastName)
                currentRace.courseName = dataFull.split("\n")[1].strip()
                raceInfo = dataFull.split("\n")[0].split("-")
                currentRace.name = raceInfo[0].strip()
                raceDateString = raceInfo[1].strip().split("/")
                currentRace.date = datetime.date(int(raceDateString[2]), int(raceDateString[0]), int(raceDateString[1]))
                courseData[currentRace.name] = currentRace
                print(currentRace.name + ": hosted at " + currentRace.courseName + " on " + str(currentRace.date))
                
print("\n")

checkFile = open("Output.txt", "w")
checkFile2 = open("Output2.txt", "w")

for person in courseData["37th Annual Greensboro XC Invitational 2021"].people:
    checkFile2.write(person.firstName + " " + person.lastName)
for courseName in courseData.keys():
    for person in courseData[courseName].people:
        name = person.firstName + " " + person.lastName
        newRacer = Racer()
        if name in people.keys():
            newRacer = people[name]
            racesParticipated = ""
            for race in newRacer.times.keys():
                racesParticipated+=race + ", "
            racesParticipated = racesParticipated + ", and " + courseName
        try:
            newRacer.times[courseName] = person.time
        except:
            newRacer.times[courseName] = "NULL"
        newRacer.firstName = person.firstName
        newRacer.lastName = person.lastName
        newRacer.school = person.school
        people[name] = newRacer
        

print(str(people["Sebastian Polge"].times))
print(str(people["Arran Swift"].times))
print(str(people["Will Soule"].times))
print(str(people["Abigail Davis"].times))
print(str(people["Jenna Pullen"].times))


wb = Workbook()
wv = Workbook()
wb2 = Workbook()
wv2 = Workbook()

peopleNew = {}
keys = list(people.keys())
keys.sort()
for key in keys: 
    peopleNew[key] = people[key]

ws1 = wb.create_sheet("Times")
ws2 = wv.create_sheet("Times")
ws3 = wb2.create_sheet("Times")
ws4 = wv2.create_sheet("Times")



i = 2
j = 2
for person in peopleNew.keys():
    if len(peopleNew[person].times.keys()) > 1:
        d1 = ws1.cell(row=j,column=1,value=person)
        d1 = ws2.cell(row=j,column=1,value=person)
        j+=1
    d1 = ws3.cell(row=i,column=1,value=person)
    d1 = ws4.cell(row=i,column=1,value=person)
    i+=1


races = {}
racesList = list(courseData.keys())
racesList.sort()
for race in racesList:
    races[race] = courseData[race]

i = 2
for raceName in races.keys():
    d1 = ws1.cell(row=1, column=i, value=raceName)
    d1 = ws2.cell(row=1, column=i, value=raceName)
    d1 = ws3.cell(row=1, column=i, value=raceName)
    d1 = ws4.cell(row=1, column=i, value=raceName)


    i+=1


rowNum = 2
rowNumDupe = 2
for person in peopleNew.values():
    for raceName in person.times.keys():
        if len(person.times.keys()) > 1:
            for i in range(2, len(courseData.keys())+2):
                value = ws1.cell(row=1, column=i).value
                if raceName == value:
                    d1 = ws1.cell(row=rowNum, column=i,value=convertToSec(person.times[raceName]))
                    d1 = ws2.cell(row=rowNum, column=i,value=convertToSec(person.times[raceName]))
        for i in range(2, len(courseData.keys())+2):
            value = ws1.cell(row=1, column=i).value
            if raceName == value:
                d1 = ws3.cell(row=rowNumDupe, column=i,value=convertToSec(person.times[raceName]))
                d1 = ws4.cell(row=rowNumDupe, column=i,value=convertToSec(person.times[raceName]))
        
    rowNumDupe+=1
    if len(person.times.keys()) > 1:
        rowNum+=1


print("\n\n")
for race in races.keys():
    print(race + ": " + str(races[race].date))

i = 2
for raceName in races.keys():
    d1 = ws1.cell(row=1, column=i, value=str(races[raceName].date))
    d1 = ws2.cell(row=1, column=i, value=str(races[raceName].date))
    d1 = ws3.cell(row=1, column=i, value=str(races[raceName].date))
    d1 = ws4.cell(row=1, column=i, value=str(races[raceName].date))

    i+=1


wb.save("XC Sheet No Dupes.xlsx")
wv.save("XC Sheet Seconds No Dupes.xlsx")
wb2.save("XC Sheet.xlsx")
wv2.save("XC Sheet Seconds.xlsx")




# To-Do:
# -Seperate out people only in one race
# -Prepare training data.

print("Completed!")