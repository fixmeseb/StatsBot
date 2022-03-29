import discord
#This imports Discord. Named thing.py because my old bots had their main files in thing.js, and I'm sentimental. 

from datetime import date
from datetime import datetime
from discord.utils import get
from openpyxl import load_workbook
from openpyxl import Workbook
from pyparsing import col

import pytz

est = pytz.timezone('US/Eastern')
utc = pytz.utc

# To run a new month: 
# - Update monthsPossible


intents = discord.Intents.all()
client = discord.Client(intents=intents)
secretSecret = False

@client.event
async def on_ready():
    
    CADiscord = client.get_guild(523962430179770369)
    starBoard = client.get_channel(785919963890319410)
    trueBoard = client.get_channel(810712965900664882)
    firingSquad = client.get_channel(814578472861171762)
    boards = {
        starBoard: "⭐",
        trueBoard: "true",
        firingSquad: "catgun",
 
    }


    validCheckers = []
    validityMessage = await CADiscord.get_channel(524453945524092929).fetch_message(904889501720121366)
    for reaction in validityMessage.reactions:
        async for reactor in reaction.users(limit=None,after=None):
            if not (reactor.id in validCheckers):
                try:
                    bing = CADiscord.get_member(reactor.id).name
                    validCheckers.append(reactor.id)
                except:
                    bing = "incorrect"
    BotRole = CADiscord.get_role(560094358033006629)
    for member in CADiscord.members:
        if BotRole in member.roles and not(member.id in validCheckers):
            validCheckers.append(member.id)
        if secretSecret == True and not(member.id in validCheckers):
            validCheckers.append(member.id)
    
    reactorsList = ""
    for valid in validCheckers:
        reactorsList+=CADiscord.get_member(valid).name
        reactorsList+=", "
    reactorsList = reactorsList[0:len(reactorsList)-2:]
    print(reactorsList)

    boardsTotals = {}
    messageNumbers = {}
    for board in boards:
        quants = {}
        for checker in validCheckers:
            quants[checker] = 0
        quants["Total"] = 0
        boardsTotals[board.name] = quants
    
    for board in boards:
        quants = {}
        for checker in validCheckers:
            quants[checker] = 0
        quants["Total"] = 0
        messageNumbers[board.name] = quants

    for channel in boards:
        async for message in channel.history(limit=None):
            if len(message.embeds) >= 1 and message.author.id == 700796664276844612 and message.content != "":
                print(message.content + " (" + str(len(message.embeds)) + ")")
                print("Check: " + str(message.embeds[0].fields[0].value))
                try:
                    messageID = int(str(message.embeds[0].fields[0].value)[len(str(message.embeds[0].fields[0].value))-19:len(str(message.embeds[0].fields[0].value))-1:])
                    print("Message ID: " + str(messageID))
                    channelID = int(str(message.embeds[0].fields[0].value)[len(str(message.embeds[0].fields[0].value))-38:len(str(message.embeds[0].fields[0].value))-20:])
                    print("Channel ID: " + str(channelID))
                except:
                    messageID = int(str(message.embeds[0].fields[0].value)[len(str(message.embeds[0].fields[0].value))-21:len(str(message.embeds[0].fields[0].value))-3:])
                    print("Message ID: " + str(messageID))
                    channelID = int(str(message.embeds[0].fields[0].value)[len(str(message.embeds[0].fields[0].value))-40:len(str(message.embeds[0].fields[0].value))-22:])
                    print("Channel ID: " + str(channelID))
                #785926058133684325
                #523966359877582850
                message = await client.get_channel(int(channelID)).get_partial_message(int(messageID)).fetch()
                #Check: [**Jump to Message**](https://discord.com/channels/523962430179770369/523966359877582850/785926058133684325)
                #Check: **[Jump to Message](https://discord.com/channels/523962430179770369/523966359877582850/856948232268808202)**
                starNumber = 0
                for reaction in message.reactions:
                    try:
                        if reaction.emoji.name == boards[channel]:
                            starNumber+=reaction.count
                    except:
                        if reaction.emoji == boards[channel]:
                            starNumber+=reaction.count
                print(message.author.name + " (" + str(starNumber) + "): " + message.content)
                try:
                    boardsTotals[channel.name][message.author.id]+=starNumber
                    messageNumbers[channel.name][message.author.id]+=1
                except:
                    valid = False
                boardsTotals[channel.name]["Total"]+=starNumber
                messageNumbers[channel.name]["Total"]+=1
                print(str(boardsTotals[channel.name]["Total"]))
                print(str(messageNumbers[channel.name]["Total"]))
                print("\n") 
    
    colNum = 2
    for board in boards:
        wb = Workbook()
        ws1 = wb.create_sheet("Star Counts")
        ws2 = wb.create_sheet("Message Counts")
        print(board.name)
        print("Totals: ")
        rowNum = 2
        for entry in boardsTotals[board.name].keys():
            try:
                ws1.cell(row=rowNum,column=1,value=client.get_user(entry).name)
                ws2.cell(row=rowNum,column=1,value=client.get_user(entry).name)
            except:
                ws1.cell(row=rowNum,column=1,value=entry)
                ws2.cell(row=rowNum,column=1,value=entry)

            try:
                if boardsTotals[board.name][entry] != 0:
                    print(client.get_user(entry).name + ": " + str(boardsTotals[board.name][entry]))
            except:
                print(entry + ": " + str(boardsTotals[board.name][entry]))
            ws1.cell(row=rowNum, column=colNum, value=boardsTotals[board.name][entry])

            try:
                if messageNumbers[board.name][entry] != 0:
                    print(client.get_user(entry).name + ": " + str(messageNumbers[board.name][entry]))
            except:
                print(entry + ": " + str(messageNumbers[board.name][entry]))  
            ws2.cell(row=rowNum,column=colNum,value=messageNumbers[board.name][entry])
            rowNum+=1          
        wb.save(board.name + " Stats.xlsx")
        #colNum+=1
    print("Completed!")




    

client.run('NjYyNzg4MTk1NzM3NDAzNDQy.Xg_Dmw.dyg4Kch4KxX6C6bDZAcx-Le2TVs')