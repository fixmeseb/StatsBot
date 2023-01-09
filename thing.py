import discord
#The old StatBot code I think. When StatBot actually answered querys and whatnot. Probably. Added this 12.12.2022. 
import xlrd
import xlwt
import pytz
from xlwt import Workbook
from xlutils.copy import copy 
from xlrd import open_workbook
from datetime import date
from datetime import datetime
from discord.utils import get
from openpyxl import load_workbook
import time
import matplotlib.pyplot as plt
from PIL import Image
import numpy as np

wb = Workbook()
shutDown = True
est = pytz.timezone('US/Eastern')
utc = pytz.utc
fmt = '%Y-%m-%d %H:%M:%S %Z%z'
optLevels = [-1, 0, 1, 2]
print("Your program has begun running.")
client = discord.Client()
def error():
    embed = discord.Embed(title="Error", description="Sorry, something went wrong.", color=0xFF9900)
    embed.set_footer(text="Created by The Invisible Man", icon_url="https://cdn.discordapp.com/avatars/366709133195476992/01cb7c2c7f2007d8b060e084ea4eb6fd.png?size=512")
    return embed
def errorLOCATE():
    embed = discord.Embed(title="Error", description="Sorry, we could not locate the requested user.", color=0xFF9900)
    embed.set_footer(text="Created by The Invisible Man", icon_url="https://cdn.discordapp.com/avatars/366709133195476992/01cb7c2c7f2007d8b060e084ea4eb6fd.png?size=512")
    return embed
def findUser(userID):
    return client.get_user(int(userID))
def getUserOptLevel(userID):
    optSheets = open("optSheets.txt", "r")
    optString = optSheets.read()
    optList = optString.split('\n')
    counter = -1
    for opt in optList:
        optFile = open(opt + ".txt", "r")
        optString = optFile.read()
        optArray = optString.split("\n")
        if str(userID) in optArray:
            print(str(userID) + "'s opt level is " + str(counter))
            return counter
        else:
            counter+=1
    return -5
@client.event
async def on_ready():
    print('We have logged in as {0.user}'.format(client))
    print("\n")

@client.event
async def on_message(message):
    if message.author == client.user:
        return
    if message.content.startswith("&optHelp"):
        print("OptHelp request made by " + message.author.name + " at " + str(message.created_at) + " on guild " + message.guild.name)
        userOptLevel = getUserOptLevel(message.author.id)
        userOpt = "Your current Opt Level is " + str(userOptLevel) + "."
        embed = discord.Embed(title="OptHelp", description=userOpt, color=0xFF9900)
        embed.add_field(name="&optLevel -1", value="No information is collected at any point on the user. Cannot use commands.", inline=False)
        embed.add_field(name="&optLevel 0", value="Anonymus Data is collected on the user. Can use anonymus commands.", inline=False)
        embed.add_field(name="&optLevel 1", value="Data is collected in spreadsheets, but only accesible by the user. Can use all commands on self, but not on other users.", inline=False)
        embed.add_field(name="&optLevel 2", value="Data is collected; accessible by anyone. Can use all commands.", inline=False)
        embed.set_footer(text="Created by The Invisible Man", icon_url="https://cdn.discordapp.com/avatars/366709133195476992/01cb7c2c7f2007d8b060e084ea4eb6fd.png?size=512")
        await message.channel.send(embed=embed)
    #Get help relating to opt levels.
    if message.content.startswith("&optLevel"):
        print("OptLevel request made by " + message.author.name + " at " + str(message.created_at) + " on guild " + message.guild.name)
        messageCont = message.content.split(" ")
        user = message.author
        if message.author.id == 366709133195476992 and len(messageCont) >= 3:
            user = client.get_user(int(messageCont[2]))
        print("Target: " + user.name + " (" + str(user.id) + ")")
        userOptLevel = getUserOptLevel(user.id)
        optLevelDesired = int(messageCont[1])
        if optLevelDesired == userOptLevel:
            await message.channel.send("You already are this level!")
        else:
            if userOptLevel != -5:
                optLevelCurrent = open("opt" + str(userOptLevel) + ".txt", "r")
                currentOptLevelString = optLevelCurrent.read()
                currentOptLevelArray = currentOptLevelString.split("\n")
                print("Booglybooglyboo")
                if str(user.id) in currentOptLevelArray:
                    currentOptLevelArray.remove(str(user.id))
                    print("Removed the id.")
                if str(user.id) in currentOptLevelArray:
                    print("The id's still there!")
                optLevelCurrent.close()
                optLevelCurrent = open("opt" + str(userOptLevel) + ".txt", "w")
                optLevelCurrentString = ""
                for item in currentOptLevelArray:
                    optLevelCurrentString = optLevelCurrentString + "\n" + item
                optLevelCurrent.write(optLevelCurrentString)
                optLevelCurrent.close()
            optLevelNew = open("opt" + str(optLevelDesired) + ".txt", "r")
            optLevelNewString = optLevelNew.read()
            optLevelNew.close()
            optLevelNew = open("opt" + str(optLevelDesired) + ".txt", "w")
            optLevelNewString = optLevelNewString + "\n" + str(user.id)
            optLevelNew.write(optLevelNewString)
            await message.channel.send("Changed <@" + str(user.id)   + ">'s opt level to " + str(optLevelDesired) + ".")
            await message.channel.send("Note: Opt Level changes will not function correctly until after the next time stat count is run.")
    #Change the opt level of the current user.
    if message.content.startswith('&userInfo') or message.content.startswith('&userinfo'):
        print("User Info request made by " + message.author.name + " at " + str(message.created_at) + " on guild " + message.guild.name)
        if shutDown == False:
            user = message.author
            userID = str(user.id)
            messageContentList = message.content.split(" ")
            if len(messageContentList) > 1:
                userID = 0
                if messageContentList[1].startswith("<@"): 
                    userPing = messageContentList[1]
                    userPingList = userPing.split('!')
                    userPingThing = userPingList[1].split('>')
                    userID = userPingThing[0]
                else:
                    userID = messageContentList[1]
                user = client.get_user(int(userID))
            
            userOptLevel = getUserOptLevel(user.id)
            print("Target username: " + user.name) 
            print("Target user ID: " + str(user.id))
                    
            messageContentList = message.content.split(" ")
            userURL = str(user.avatar_url)
            print("Target user URL: " + userURL)
            if (userOptLevel >= 2 and getUserOptLevel(message.author.id)) or (userOptLevel == 1 and user.id == message.author.id):
                wz = xlrd.open_workbook("C:\\Users\\Sebastian_Polge\\OneDrive-CaryAcademy\\Documents\\meNewBot\\Verity\\StatsBot\\Complete.xls")
                statBook = xlrd.open_workbook("C:\\Users\\Sebastian_Polge\\OneDrive-CaryAcademy\\Documents\\meNewBot\\Verity\\StatsBot\\TimesMan.xls")
                statSheet = statBook.sheet_by_index(0)
                userRow = -1
                mostActiveHour = "-1"
                mostActiveHourNumber = 0
                for i in range(statSheet.nrows):
                    if i != 0 and statSheet.cell_value(i, 0) == str(user.id):
                        userRow = i
                if userRow == -1:
                    errorLOCATE()
                else:
                    for i in range(statSheet.ncols):
                        if i != 0:
                            if int(statSheet.cell_value(userRow, i)) >= mostActiveHourNumber and int(statSheet.cell_value(userRow, i)) != 0:
                                mostActiveHour = str(statSheet.cell_value(0, i))
                                mostActiveHourNumber = int(statSheet.cell_value(userRow, i))
                sheet = wz.sheet_by_index(0)
                sheetList = wz.sheets()
                string = message.guild.name
                numberOfChannels = 0
                positionOfUser = 0                
                guildSheet = wz.sheet_by_name(string)
                numberOfChannels = guildSheet.ncols - 1
                actionCompleted = False
                for i in range(guildSheet.nrows):             
                    if str(guildSheet.cell_value(i, 0)).isdigit() == True and int(guildSheet.cell_value(i, 0)) == user.id:
                        positionOfUser = i
                        totalMessagesSentServer = guildSheet.cell_value(positionOfUser, numberOfChannels - 3)
                        messageAverageOnServer = guildSheet.cell_value(positionOfUser, numberOfChannels)
                        activeChannel = "No channels had more than 10 messages sent in them."
                        activeChannelNumber = -1
                        for v in range(guildSheet.ncols):
                            if v != 0 and v < guildSheet.ncols - 5:
                                if int(guildSheet.cell_value(positionOfUser, v)) > activeChannelNumber and int(guildSheet.cell_value(positionOfUser, v)) > 10:
                                    activeChannel = "#" + guildSheet.cell_value(0, v)
                                    activeChannelNumber = int(guildSheet.cell_value(positionOfUser, v)) 
                        activeChannelA = ""
                        activeChannelNumberA = 0
                        for v in range(guildSheet.ncols):
                            channel = message.channel
                            for z in message.guild.text_channels:
                                if (guildSheet.cell_value(positionOfUser, v) == z.name):
                                    channel = z
                            currentDate = datetime.date(datetime.now())
                            channelAge = channel.created_at
                            delta = currentDate - datetime.date(channelAge)
                            if v != 0 and v < guildSheet.ncols - 5:
                                sporg = int(guildSheet.cell_value(positionOfUser, v)) / delta.days
                                if activeChannelNumberA < sporg:
                                    activeChannelNumberA = sporg
                                    activeChannelA = "#" + guildSheet.cell_value(0, v)
                        totalMessagesSent = 0
                        messageAverageServer = ""
                        messageAverage = 0
                        for server in sheetList:
                            serverMembers = []
                            for i in range(server.nrows): 
                                beep = server.cell_value(i, 0)
                                serverMembers.append(beep)
                            if str(user.id) in serverMembers:
                                locationUser = serverMembers.index(str(user.id))
                                totalMessageLocation = server.ncols - 4
                                averageMessageLocation = server.ncols - 1
                                print("Location User: " + str(locationUser))
                                print("Total Message Location: " + str(totalMessageLocation))
                                totalMessagesSent = totalMessagesSent + int(server.cell_value(locationUser, totalMessageLocation))
                                if messageAverage < server.cell_value(locationUser, averageMessageLocation) and server.cell_value(locationUser, averageMessageLocation) > 0:
                                    messageAverage = server.cell_value(locationUser, averageMessageLocation)
                                    messageAverageServer = server.name
                                #print (str(messageAverage) + " on " + messageAverageServer)
                        embed = discord.Embed(title=user.name, color=0xFF9900)
                        if message.guild.name == "The CA Discord" and int(guildSheet.cell_value(positionOfUser, 12)) >= 1000:
                            embed.add_field(name="Messages Sent in Counting and Recursion: ", value=int(guildSheet.cell_value(positionOfUser, 12)), inline=False)
                            embed.add_field(name="Messages Sent Not in Counting and Recursion: ", value=int(totalMessagesSentServer) - int(guildSheet.cell_value(positionOfUser, 12)), inline=False)
                        else:
                            embed.add_field(name="Total Messages Sent on this Server: ", value=int(totalMessagesSentServer), inline=False)
                        embed.add_field(name="Highest Message Number Channel on this Server: ", value=activeChannel, inline=False)
                        embed.add_field(name="Most Active Channel on this Server: ", value=str(activeChannelA), inline=False)
                        embed.add_field(name="Average Messages on this Server Per Day: ", value=str(round(messageAverageOnServer, 2)), inline=False)
                        embed.add_field(name="Most Active Server: ", value=messageAverageServer, inline=False)
                        embed.add_field(name="Total Messages Sent: ", value=int(totalMessagesSent), inline=False)
                        embed.add_field(name="Most Active Time: ", value=mostActiveHour, inline=False)
                        embed.set_thumbnail(url=userURL)
                        embed.set_footer(text="Created by The Invisible Man", icon_url="https://cdn.discordapp.com/avatars/366709133195476992/01cb7c2c7f2007d8b060e084ea4eb6fd.png?size=512")
                        await message.channel.send(embed=embed)
                        print("Completed\n")
                        actionCompleted = True
                if actionCompleted == False:
                    embed = error()
                    await message.channel.send(embed=embed)
            else:
                await message.channel.send("You do not have the required clearance to use this command.")
        else:
            await message.channel.send("Sorry for the inconvience, but StatBot is currently shut down with most commands so people have the chance to opt in to the bot.")
        #Gets user info.
    #The user info command
    if message.content.startswith('&serverInfo') or message.content.startswith("&serverinfo"):
        userOptLevel = getUserOptLevel(message.author.id)
        print("Server Info request made by " + message.author.name + " at " + str(message.created_at) + " for guild " + message.guild.name)
        if shutDown == False:
            if userOptLevel >= 0:
                wz = xlrd.open_workbook("C:\\Users\\Sebastian_Polge\\OneDrive-CaryAcademy\\Documents\\meNewBot\\Verity\\StatsBot\\Complete.xls")
                guildSheet = wz.sheet_by_name(message.guild.name)
                positionOfTotalColumn = guildSheet.ncols - 4
                positionOfAverageMessage = guildSheet.ncols -1
                activeMember = ""
                activeMemberActivity = 0
                activeMemberNoCount = ""
                activeMemberNoCountNum = 0
                totalMessages = 0
                for i in range(guildSheet.nrows):
                    if i != 0 and i < guildSheet.nrows-1:
                        print("i/total col position = " + str(i) + "/" + str(positionOfTotalColumn))
                        totalMessages += int(guildSheet.cell_value(i, positionOfTotalColumn))
                        if (guildSheet.cell_value(i, positionOfAverageMessage) > activeMemberActivity):
                            activeMemberActivity = guildSheet.cell_value(i, positionOfAverageMessage)
                            activeMemberID = guildSheet.cell_value(i, 0)
                            activeMember = message.guild.get_member(int(activeMemberID))
                for i in range(guildSheet.nrows):
                    if (i != 0) and message.guild.name == "The CA Discord":
                        boop = guildSheet.cell_value(i, guildSheet.ncols - 4) - guildSheet.cell_value(i, 12)
                        beep = boop/guildSheet.cell_value(i, guildSheet.ncols - 2)
                        if (beep > activeMemberNoCountNum):
                            activeMemberNoCountNum = beep
                            activeMemberNoCountID = guildSheet.cell_value(i, 0)
                            activeMemberNoCount = message.guild.get_member(int(activeMemberNoCountID))
                totalPerChannel = 0
                sumTotal = 0
                channelName = ""
                countingAndRecursionInt = -1
                endCol = guildSheet.ncols - 4
                for i in range(guildSheet.ncols):
                    sumTotal = 0
                    if (i != 0 and i < endCol):
                        for v in range(guildSheet.nrows - 1):
                            if (v != 0):
                                sumTotal+=int(guildSheet.cell_value(v, i))
                            if (sumTotal > totalPerChannel):
                                if guildSheet.cell_value(0, i) != "counting-and-recursion":
                                    channelName = guildSheet.cell_value(0, i)
                                    totalPerChannel = sumTotal
                                else:
                                    countingAndRecursionInt = sumTotal
                #print (channelName +": " + str(totalPerChannel))
                memberCount = guildSheet.nrows - 1
                channelCount = guildSheet.ncols - 5
                owner = message.guild.owner.name
                currentDate = datetime.date(datetime.now())
                serverAge = message.guild.created_at
                delta = currentDate - datetime.date(serverAge)
                messageActivity = int(totalMessages) / delta.days
                embed = discord.Embed(title=message.guild.name, color=0xFF9900)
                embed.add_field(name="Total Messages Sent on this Server: ", value = int(totalMessages), inline=False)
                embed.add_field(name="Most Active Member: ", value=activeMember.name, inline=False)
                if message.guild.name == "The CA Discord":
                    embed.add_field(name="Most Active Member, Not Including #counting-and-recursion: ", value=activeMemberNoCount.name, inline=False)
                if (countingAndRecursionInt > totalPerChannel):
                    embed.add_field(name="Highest Message Count per Channel, Not Including #counting-and-recursion: ", value="#" + channelName, inline=False)
                    embed.add_field(name="Highest Message Count per Channel", value="#counting-and-recursion: ", inline=False)
                else:
                    embed.add_field(name="Highest Message Count per Channel: ", value="#" + channelName, inline=False)
                embed.add_field(name="Number of Members: ", value=memberCount, inline=False)
                embed.add_field(name="Number of Text Channels: ", value=channelCount, inline=False)
                embed.add_field(name="Message Activity: ", value=str(round(messageActivity,2)) + " messages per day.", inline=False)
                embed.add_field(name="Owner: ", value=owner, inline=False)
                embed.set_thumbnail(url=message.guild.icon_url)
                embed.set_footer(text="Created by The Invisible Man", icon_url="https://cdn.discordapp.com/avatars/366709133195476992/01cb7c2c7f2007d8b060e084ea4eb6fd.png?size=512")
                await message.channel.send(embed=embed) 
                print("\n")
            else:
                await message.channel.send("You do not have the required clearance to use this command.")
        else:
            await message.channel.send("Sorry for the inconvience, but StatBot is currently shut down with most commands so people have the chance to opt in to the bot.")
    #Gets the info on the server.
    if message.content.startswith('&serverStatCount') and message.author.id == 366709133195476992:
        dmChannel = await message.author.create_dm()
        optAllowed = []
        anonAllowed = []
        guildCount = 1
        otherGuildCount = 1
        guildTotal = len(client.guilds)
        serverActive = message.guild
        wb = Workbook()
        sheet1 = wb.add_sheet('Sheet 1')
        i = 1
        zed = 0
        if serverActive.name == "Starserver":
            zed-=3
        memberList = serverActive.members
        memberQuant = len(memberList)
        for member in memberList:
            userOptLevel = getUserOptLevel(member.id)
            if userOptLevel == -5 and member.bot == False:
                print("Added " + member.name + " to default optLevel.")
                defaultOpt = open("opt0.txt", "a")
                defaultOpt.write("\n" + str(member.id))
                userOptLevel = 0
            if member.bot == True and userOptLevel == -5:
                print("Added " + member.name + " to default optLevel for bots.")
                defaultOpt = open("opt2.txt", "a")
                defaultOpt.write("\n" + str(member.id))
                userOptLevel = 2
            if userOptLevel >= 0:
                anonAllowed.append(member.id)
            if userOptLevel >= 1:
                optAllowed.append(member.id)
                sheet1.write(i, 0, str(member.id))
                i+=1
                print("Added " + member.name + "(" + str(member.id) + ") (opt level " + str(userOptLevel) + ").")
            else:
                print(member.name + "(" + str(member.id) + ") has an optLevel of " + str(userOptLevel))
            wb.save(serverActive.name + ".xls")
        channelList = serverActive.text_channels
        channelQuant = len(channelList)
        x = 1
        for channel in channelList:
            if channel.name != "robot-game" and channel.name != "starfall-private-space" and channel.name != "ca-nerd-squad" and channel.name != "vent":
                sheet1.write(0, x, channel.name)
                x+=1
                print("Added #" + channel.name)
        wb.save(serverActive.name + ".xls")
        server = serverActive.text_channels  
        channelNumberIc = 1 
        for channel in server:
            if channel.name != "robot-game" and channel.name != "starfall-private-space" and channel.name != "ca-nerd-squad" and channel.name != "vent":
                y = 1
                authorMessageQuant = {
                }
                for member in memberList:
                    authorMessageQuant[str(member.id)] = 0
                print(str(channelNumberIc) + "/" + str(channelQuant) + "; " + str(otherGuildCount) + "/" + str(guildTotal) + " #" + channel.name)
                async for message in channel.history(limit=None):
                    for member in memberList:
                        if member.name == message.author.name:
                            authorMessageQuant[str(member.id)] = str(int(authorMessageQuant[str(member.id)])+ 1)
                wr = xlrd.open_workbook("C:\\Users\\Sebastian_Polge\\OneDrive-CaryAcademy\\Documents\\meNewBot\\Verity\\StatsBot\\" + serverActive.name + ".xls") 
                sheet = wr.sheet_by_index(0) 
                for stuff in range(sheet.ncols):
                    if sheet.cell_value(0,stuff) == channel.name:
                        print("Channel is found!") 
                        y = stuff
                anonTotal = 0
                for member in authorMessageQuant:
                    x = 1
                    for mStuff in range(sheet.nrows):
                        if sheet.cell_value(mStuff, 0) == member:
                            print("Author is found! " + sheet.cell_value(mStuff, 0) + "/" + member) 
                            x = mStuff
                    messageCount = authorMessageQuant[member]
                    for i in optAllowed:
                        print(str(member) + "/" + str(i))
                    if int(member) in optAllowed:
                        print(sheet.cell_value(x, 0) + " opted in")
                        print(str(x) + "/" + str(memberQuant) + "; " + str(y) + "/" + str(channelQuant) + "; " + str(guildCount) + "/" + str(guildTotal) + " #" + str(channel.name))
                        sheet1.write(x, y, int(messageCount))
                        zed = y + 1
                    if int(member) in anonAllowed:
                        anonTotal+=int(messageCount)
                    s = message.guild.name
                    if s == "Cary Academy D&D- A Band of Fools" or s == "Cary Academy D&D- A Band of Fools":
                        s = "Cary Academy DnD"
                    wb.save(s + ".xls")
            channelNumberIc+=1
            sheet1.write(x+1,y,anonTotal)
            print("Completed channel #" + channel.name)
            
        x = 1
        print("Location of First Total Column: " + str(zed))
        sheet1.write(0, zed, "Message Total:")
        sheet1.write(0, zed + 1, "Date Joined:")
        sheet1.write(0, zed + 2, "Days on Server:")
        sheet1.write(0, zed + 3, "Message Average:")
        for member in memberList:
            if member.id in optAllowed:
                string = ""
                n = y
                while n > 0:
                    n, remainder = divmod(n - 1, 26)
                    string = chr(65 + remainder) + string
                value1 = "B" + str(x + 1)
                value2 = string + str(x + 1)
                sheet1.write(x, zed, xlwt.Formula("SUM(" + value1 + ":" + value2 + ")"))
                dateTimeFull = member.joined_at
                dateTimeSplit = str(dateTimeFull).split()
                sheet1.write(x, zed + 1, dateTimeSplit[0])
                dateTimeFull = member.joined_at
                dateJoined = datetime.date(dateTimeFull)
                dateTimeSplit = str(dateTimeFull).split()
                currentDate = datetime.date(datetime.now())
                delta = currentDate - dateJoined
                sheet1.write(x, zed + 2, delta.days)
                sheet1.write(x, zed + 3, xlwt.Formula("SUM(" + value1 + ":" + value2 + ")/" + str(delta.days))) 
                x+=1
        s = serverActive.name
        if s == "Cary Academy D&D- A Band of Fools" or s == "Cary Academy D&D- A Band of Fools":
            s = "Cary Academy DnD"
        wb.save(s + ".xls")
        print("Complete!")
        guildCount+=1
        otherGuildCount +=1
        await dmChannel.send("Completed " + serverActive.name)
    #Counts the stats of one specific server.
    if message.content.startswith("&serverActiveList") or message.content.startswith("&serveractivelist"):
        print("Server Info request made by " + message.author.name + " at " + str(message.created_at) + " for guild " + message.guild.name)
        userOptLevel = getUserOptLevel(message.author.id)
        if shutDown == False:
            if userOptLevel >= 2:    
                wz = xlrd.open_workbook("C:\\Users\\Sebastian_Polge\\OneDrive-CaryAcademy\\Documents\\meNewBot\\Verity\\StatsBot\\Complete.xls")
                guildSheet = wz.sheet_by_name(message.guild.name)
                if len(message.content.split()) > 1:
                    uncludeChannels = []
                    for i in message.content.split():
                        if message.content.split().index(i) != 0:
                            uncludeChannels.append(i)
                            print(i)
                    positionOfAverageColumn = guildSheet.ncols - 1
                    positionOfDaysCol = guildSheet.ncols - 2
                    positionOfTotalCol = guildSheet.ncols - 4
                    embed = discord.Embed(title="Most Active On Server", color=0xFF9900)
                    x = []
                    for i in range(guildSheet.nrows):
                        if i != 0:
                            grandTotal = 0
                            for c in range(guildSheet.ncols):
                                if c > 0 and c < positionOfTotalCol:
                                    if not guildSheet.cell_value(0, c) in uncludeChannels:
                                        grandTotal += guildSheet.cell_value(i, c)
                                        y = [guildSheet.cell_value(i, 0), guildSheet.cell_value(i, positionOfAverageColumn)]
                                    x.append(y)
                    for i in sorted(x, key = lambda x: x[1])[::-1]:
                        messageTotal = z
                        bloop = round(i[1], 2)
                        userUserName = message.guild.get_member(i[0])

                        embed.add_field(name=userUserName.name, value= "Average of " + str(bloop) + " messages per day (" + str(), inline=False)
                        #print(str(i[0]) + " added")
                        #print("X len: " + str(len(x)))

                    embed.set_thumbnail(url=message.guild.icon_url)
                    embed.set_footer(text="Created by The Invisible Man", icon_url="https://cdn.discordapp.com/avatars/366709133195476992/01cb7c2c7f2007d8b060e084ea4eb6fd.png?size=512")
                    await message.channel.send(embed=embed)

                else:
                    positionOfAverageColumn = guildSheet.ncols - 1
                    embed = discord.Embed(title="Most Active On Server", color=0xFF9900)
                    x = []
                    for i in range(guildSheet.nrows):
                        if i != 0:
                            y = [guildSheet.cell_value(i, 0), guildSheet.cell_value(i, positionOfAverageColumn)]
                            x.append(y)
                    
                    for i in sorted(x, key = lambda x: x[1])[::-1]:
                        bloop = round(i[1], 2)
                        print('"' + str(i[0]) + '"')
                        userUserName = message.guild.get_member(int(i[0]))

                        embed.add_field(name=userUserName.name, value= "Average of " + str(bloop) + " messages per day", inline=False)
                        #print(str(i[0]) + " added")
                    embed.set_thumbnail(url=message.guild.icon_url)
                    embed.set_footer(text="Created by The Invisible Man", icon_url="https://cdn.discordapp.com/avatars/366709133195476992/01cb7c2c7f2007d8b060e084ea4eb6fd.png?size=512")
                    await message.channel.send(embed=embed)
                print("\n")
            else:
                await message.channel.send("You do not have the required clearance to use this command.")
        else:
            await message.channel.send("Sorry for the inconvience, but StatBot is currently shut down with most commands so people have the chance to opt in to the bot.")
    #Gets the list of the most active members on a server.
    if message.content.startswith("&channelInfo") or message.content.startswith("&channelinfo"):
        print("Channel Info request made by " + message.author.name + " at " + str(message.created_at) + " on guild " + message.guild.name)
        userOptInfo = getUserOptLevel(message.author.id)
        if shutDown == False:
            if userOptInfo >= 0: 
                channel = message.channel
                if len(message.content.split(" ")) > 1:
                    messageSplit = message.content.split(" ")
                    channelNotif = messageSplit[1].split("#")
                    channelNotifSplit2 = channelNotif[1].split(">")
                    channelID = channelNotifSplit2[0]
                    channel = message.guild.get_channel(int(channelID))
                    
                wz = xlrd.open_workbook("C:\\Users\\Sebastian_Polge\\OneDrive-CaryAcademy\\Documents\\meNewBot\\Verity\\StatsBot\\Complete.xls")
                guildSheet = wz.sheet_by_name(message.guild.name)
                channelPos = -1
                for i in range(guildSheet.ncols):
                    if guildSheet.cell_value(0, i) == channel.name:
                        channelPos = i
                messageHighID = ""
                messageHigh = 0
                totalMessages = 0
                print(channel.name + "/" + guildSheet.cell_value(0, channelPos))
                if channelPos != 0 and channelPos != -1:
                    for i in range(guildSheet.nrows - 1):
                        if i != 0:
                            totalMessages = totalMessages + int(guildSheet.cell_value(i,channelPos))
                            if int(guildSheet.cell_value(i, channelPos)) > messageHigh and getUserOptLevel(int(guildSheet.cell_value(i, 0))) >= 2:
                                messageHigh = guildSheet.cell_value(i, channelPos)
                                messageHighID = guildSheet.cell_value(i, 0)
                            else:
                                if int(guildSheet.cell_value(i, channelPos)) == messageHigh:
                                    messageHighID = messageHighID + "/" + guildSheet.cell_value(i, 0)

                currentDate = datetime.date(datetime.now())
                channelAge = channel.created_at
                delta = currentDate - datetime.date(channelAge)
                messageActivity = totalMessages / delta.days
                messageHighUser = channel.guild.get_member(int(messageHighID))
                embed = discord.Embed(title="#" + channel.name + " info", color=0xFF9900)
                if userOptInfo >= 2:
                    embed.add_field(name="Person with highest message count: ", value=messageHighUser.name + " (" + str(int(messageHigh)) + " messages)", inline=False)
                embed.add_field(name="Total Messages Sent:", value=totalMessages, inline=False)
                embed.add_field(name="Message Activity: ", value=str(round(messageActivity, 2)) + " per day", inline=False)
                embed.add_field(name="Server: ", value=channel.guild, inline=False)
                embed.set_thumbnail(url=message.guild.icon_url)
                embed.set_footer(text="Created by The Invisible Man", icon_url="https://cdn.discordapp.com/avatars/366709133195476992/01cb7c2c7f2007d8b060e084ea4eb6fd.png?size=512")
                await message.channel.send(embed=embed)
            else:
                await message.channel.send("You do not have the required clearance to use this command.")

        else:
            await message.channel.send("Sorry for the inconvience, but StatBot is currently shut down with most commands so people have the chance to opt in to the bot.")
    #Channel Info. 
    if message.content.startswith("&help"):
        print("Help request made by " + message.author.name + " at " + str(message.created_at) + " on guild " + message.guild.name)
        embed = discord.Embed(title="Help")
        if shutDown == False:
            embed = discord.Embed(title="Help", description="Hello, I'm StatBot! Here are some of my functions: ", color=0xFF9900)
            embed.add_field(name="&userInfo", value="Get information on a user. Either use the user's id or ping the user.", inline=False)
            embed.add_field(name="&serverInfo", value="Get information on the server.", inline=False)
            embed.add_field(name="&serverActiveList", value="Get a list of the most active members on the server. Uses total messages sent over time on server.", inline=False)
            embed.add_field(name="&channelInfo", value='Get info on a specific channel. Use the channel' + "'" + 's "ping" to get the channel, or add no other arguements to get the current channel' + "'" + "s info.", inline=False)
            embed.add_field(name="&timeCounter", value='Get activity times on a specific channel. Use the channel' + "'" + 's "ping" to get the channel. YOU WILL NEED TO GIVE IT SOME TIME!', inline=False)
            embed.add_field(name="&userChart", value='Get a bar graph with the activity times of the user. Either ping the user or use their id.', inline=False)
            embed.add_field(name="&optHelp", value="Display the various options for opt level.", inline=False)
            embed.add_field(name="&optLevel", value="Changes the optLevel to the specified value after the command. Use `&optHelp` to see the different options.")
            embed.set_footer(text="Created by The Invisible Man", icon_url="https://cdn.discordapp.com/avatars/366709133195476992/01cb7c2c7f2007d8b060e084ea4eb6fd.png?size=512")
        else:
            embed = discord.Embed(title="Help", description="StatBot is temporarily shut down so people hve the chance to opt in. The two current commands available are:", color=0xFF9900)
            embed.add_field(name="&optHelp", value="Display the various options for opt level.", inline=False)
            embed.add_field(name="&optLevel", value="Changes the optLevel to the specified value after the command. Use `&optHelp` to see the different options.")
            embed.set_footer(text="Created by The Invisible Man", icon_url="https://cdn.discordapp.com/avatars/366709133195476992/01cb7c2c7f2007d8b060e084ea4eb6fd.png?size=512")
        await message.channel.send(embed=embed)
    #Normal Help message for all users.
    if message.content.startswith("&helpSpecial") and message.author.id == 366709133195476992:
        print("Special help request made by " + message.author.name + " at " + str(message.created_at) + " on guild " + message.guild.name)
        embed = discord.Embed(title="Special Help", description="Hello, I'm StatBot! Here are some of my functions: ", color=0xFF9900)
        embed.add_field(name="&userInfo", value="Get information on a user. Either use the user's id or ping the user.", inline=False)
        embed.add_field(name="&serverInfo", value="Get information on the server.", inline=False)
        embed.add_field(name="&serverActiveList", value="Get a list of the most active members on the server. Uses total messages sent over time on server.", inline=False)
        embed.add_field(name="&channelInfo", value='Get info on a specific channel. Use the channel' + "'" + 's "ping" to get the channel, or add no other arguements to get the current channel' + "'" + "s info.", inline=False)
        embed.add_field(name="&timeCounter", value='Get activity times on a specific channel. Use the channel' + "'" + 's "ping" to get the channel. YOU WILL NEED TO GIVE IT SOME TIME!', inline=False)
        embed.add_field(name="&userChart", value='Get a bar graph with the activity times of the user. Either ping the user or use their id.', inline=False)
        embed.add_field(name="&statCountAllServer", value='Count the server message stats on each server into seperate excel sheets. (Invisible Man Only)', inline=False)
        embed.add_field(name="&updateComplete", value='Combine all those multiple pesky server excel sheets into Complete.xls! (Invisible Man Only)', inline=False)
        embed.add_field(name="&timeCountAll", value='Get activity times on every server! (Invisible Man Only)', inline=False)
        embed.set_footer(text="Created by The Invisible Man", icon_url="https://cdn.discordapp.com/avatars/366709133195476992/01cb7c2c7f2007d8b060e084ea4eb6fd.png?size=512")
        await message.channel.send(embed=embed)
    #My special help. 
    if message.content.startswith('&statCountAllServer') and message.author.id == 366709133195476992:
        dmChannel = await message.author.create_dm()
        optAllowed = []
        anonAllowed = []
        guildCount = 1
        otherGuildCount = 1
        foo = True
        guildTotal = len(client.guilds)
        for serverActive in client.guilds:
            wb = Workbook()
            sheet1 = wb.add_sheet('Sheet 1')
            i = 1
            zed = 0
            if serverActive.name == "Starserver":
                zed-=3
            memberList = serverActive.members
            memberQuant = len(memberList)
            for member in memberList:
                userOptLevel = getUserOptLevel(member.id)
                if userOptLevel == -5 and member.bot == False:
                    print("Added " + member.name + " to default optLevel.")
                    defaultOpt = open("opt0.txt", "a")
                    defaultOpt.write("\n" + str(member.id))
                    userOptLevel = 0
                if member.bot == True and userOptLevel == -5:
                    print("Added " + member.name + " to default optLevel for bots.")
                    defaultOpt = open("opt2.txt", "a")
                    defaultOpt.write("\n" + str(member.id))
                    userOptLevel = 2    
                if userOptLevel >= 0:
                    anonAllowed.append(member.id)
                if userOptLevel >= 1:
                    optAllowed.append(member.id)
                    sheet1.write(i, 0, str(member.id))
                    i+=1
                    print("Added " + member.name + "(" + str(member.id) + ") (opt level " + str(userOptLevel) + ").")
                else:
                    print(member.name + "(" + str(member.id) + ") has an optLevel of " + str(userOptLevel))
                wb.save(serverActive.name + ".xls")
            channelList = serverActive.text_channels
            channelQuant = len(channelList)
            x = 1
            for channel in channelList:
                if channel.name != "robot-game" and channel.name != "starfall-private-space" and channel.name != "ca-nerd-squad" and channel.name != "vent":
                    sheet1.write(0, x, channel.name)
                    x+=1
                    print("Added #" + channel.name)
            wb.save(serverActive.name + ".xls")
            server = serverActive.text_channels  
            channelNumberIc = 1 
            for channel in server:
                if channel.name != "robot-game" and channel.name != "starfall-private-space" and channel.name != "ca-nerd-squad" and channel.name != "vent":
                    y = 1
                    authorMessageQuant = {
                    }
                    for member in memberList:
                        authorMessageQuant[str(member.id)] = 0
                    print(str(channelNumberIc) + "/" + str(channelQuant) + "; " + str(otherGuildCount) + "/" + str(guildTotal) + " #" + channel.name)
                    async for message in channel.history(limit=None):
                        for member in memberList:
                            if member.name == message.author.name:
                                authorMessageQuant[str(member.id)] = str(int(authorMessageQuant[str(member.id)])+ 1)
                    wr = xlrd.open_workbook("C:\\Users\\Sebastian_Polge\\OneDrive-CaryAcademy\\Documents\\meNewBot\\Verity\\StatsBot\\" + serverActive.name + ".xls") 
                    sheet = wr.sheet_by_index(0) 
                    for stuff in range(sheet.ncols):
                        if sheet.cell_value(0,stuff) == channel.name:
                            print("Channel is found!") 
                            y = stuff
                    anonTotal = 0
                    for member in authorMessageQuant:
                        x = 1
                        for mStuff in range(sheet.nrows):
                            if sheet.cell_value(mStuff, 0) == member:
                                print("Author is found! " + sheet.cell_value(mStuff, 0) + "/" + member) 
                                x = mStuff
                        messageCount = authorMessageQuant[member]
                        for i in optAllowed:
                            print(str(member) + "/" + str(i))
                        if int(member) in optAllowed:
                            print(sheet.cell_value(x, 0) + " opted in")
                            print(str(x) + "/" + str(memberQuant) + "; " + str(y) + "/" + str(channelQuant) + "; " + str(guildCount) + "/" + str(guildTotal) + " #" + str(channel.name))
                            sheet1.write(x, y, int(messageCount))
                            zed = y + 1
                        if int(member) in anonAllowed:
                            anonTotal+=int(messageCount)
                        s = message.guild.name
                        if s == "Cary Academy D&D- A Band of Fools" or s == "Cary Academy D&D- A Band of Fools":
                            s = "Cary Academy DnD"
                        wb.save(s + ".xls")
                channelNumberIc+=1
                print("X: " + str(x+1) + ", Y: " + str(y) + ", Value: " + str(anonTotal))
                
                if anonTotal == 2529 and foo == True:
                    foo = False
                    sheet1.write(x+1, y, anonTotal)
                print("Completed channel #" + channel.name)
                
            x = 1
            print("Location of First Total Column: " + str(zed))
            sheet1.write(0, zed, "Message Total:")
            sheet1.write(0, zed + 1, "Date Joined:")
            sheet1.write(0, zed + 2, "Days on Server:")
            sheet1.write(0, zed + 3, "Message Average:")
            for member in memberList:
                if member.id in optAllowed:
                    string = ""
                    n = y
                    while n > 0:
                        n, remainder = divmod(n - 1, 26)
                        string = chr(65 + remainder) + string
                    value1 = "B" + str(x + 1)
                    value2 = string + str(x + 1)
                    sheet1.write(x, zed, xlwt.Formula("SUM(" + value1 + ":" + value2 + ")"))
                    dateTimeFull = member.joined_at
                    dateTimeSplit = str(dateTimeFull).split()
                    sheet1.write(x, zed + 1, dateTimeSplit[0])
                    dateTimeFull = member.joined_at
                    dateJoined = datetime.date(dateTimeFull)
                    dateTimeSplit = str(dateTimeFull).split()
                    currentDate = datetime.date(datetime.now())
                    delta = currentDate - dateJoined
                    sheet1.write(x, zed + 2, delta.days)
                    sheet1.write(x, zed + 3, xlwt.Formula("SUM(" + value1 + ":" + value2 + ")/" + str(delta.days))) 
                    x+=1
            s = serverActive.name
            if s == "Cary Academy D&D- A Band of Fools" or s == "Cary Academy D&D- A Band of Fools":
                s = "Cary Academy DnD"
            wb.save(s + ".xls")
            print("Complete!")
            guildCount+=1
            otherGuildCount +=1
            await dmChannel.send("Completed " + serverActive.name)
    #Counts the stats on each server.
    if message.content.startswith('&updateComplete') and message.author.id == 366709133195476992:
        wc = Workbook()
        for serverActive in client.guilds:
            rb = xlrd.open_workbook("C:\\Users\\Sebastian_Polge\\OneDrive-CaryAcademy\\Documents\\meNewBot\\Verity\\StatsBot\\Complete.xls") 
            s = serverActive.name
            print("WorkBook Name: " + '"' + serverActive.name + '"')
            if s == "Cary Academy D&D- A Band of Fools" or s == "Cary Academy D&D- A Band of Fools":
                s = "Cary Academy DnD"
            serverStatBook = xlrd.open_workbook("C:\\Users\\Sebastian_Polge\\OneDrive-CaryAcademy\\Documents\\meNewBot\\Verity\\StatsBot\\" + s + ".xls") 
            serverStat = serverStatBook.sheet_by_index(0)
            sheetOne = wc.add_sheet(s)
            for i in range(serverStat.ncols):
                for j in range(serverStat.nrows):
                    if j == 0 or (i != serverStat.ncols - 1 and i != serverStat.ncols - 4):
                        sheetOne.write(j, i, serverStat.cell_value(j,i))
            for j in range(serverStat.nrows):
                if serverStat.cell_value(j,0) == "":
                        print("ID is empty. ")
                else:
                    if j != 0:
                        string = ""
                        n = serverStat.ncols - 5
                        while n > 0:
                            n, remainder = divmod(n - 1, 26)
                            string = chr(65 + remainder) + string
                        value1 = "B" + str(j + 1)
                        value2 = string + str(j + 1)
                        formula = "SUM(" + value1 + ":" + value2 + ")"
                        sheetOne.write(j, serverStat.ncols - 4, xlwt.Formula(formula))
                        user = client.get_user(int(serverStat.cell_value(j,0)))
                        
                        print(user.name + " (" + str(user.id) + ")")
                        member = serverActive.get_member(int(serverStat.cell_value(j,0)))
                        dateTimeFull = member.joined_at
                        dateJoined = datetime.date(dateTimeFull)
                        dateTimeSplit = str(dateTimeFull).split()
                        currentDate = datetime.date(datetime.now())
                        delta = currentDate - dateJoined
                        sheetOne.write(j, serverStat.ncols - 1, xlwt.Formula("SUM(" + value1 + ":" + value2 + ")/" + str(delta.days)))
                        wc.save("Complete.xls")
        print("Completed.")
        dmChannel = await message.author.create_dm()
        await dmChannel.send("Updates Complete.")
    #Update Complete.xls with the info from invdividual spreadsheets. 
    if message.content.startswith('&timeCounter'):
        print("Channel Activity Times request made by " + message.author.name + " at " + str(message.created_at) + " on guild " + message.guild.name)
        if shutDown == False:
            if getUserOptLevel(message.author.id) >= 0:
                oneSec = await message.channel.send("One second, please...")
                time = {
                    0:0,
                    1:0,
                    2:0,
                    3:0,
                    4:0,
                    5:0,
                    6:0,
                    7:0,
                    8:0,
                    9:0,
                    10:0,
                    11:0,
                    12:0,
                    13:0,
                    14:0,
                    15:0,
                    16:0,
                    17:0,
                    18:0,
                    19:0,
                    20:0,
                    21:0,
                    22:0,
                    23:0,
                }
                if len(message.content.split(' ')) == 1 and message.author.id == 366709133195476992: 
                    for i in message.guild.text_channels:
                        async for mess in i.history(limit=None):
                            messageTimeUTC = mess.created_at
                            source_time_zone = pytz.timezone("Etc/UTC")
                            source_date_with_timezone = source_time_zone.localize(messageTimeUTC)
                            target_time_zone = est
                            target_date_with_timezone = source_date_with_timezone.astimezone(target_time_zone)
                            print(message.channel.name)
                            hour = int(target_date_with_timezone.hour)
                            time[hour] = int(time[hour]) + 1
                    
                    listHours = []
                    embed = discord.Embed(title="Server Activity Times", color=0xFF9900)
                    for i in time:
                            embed.add_field(name=str(i) + ":00", value=str(time[i]), inline=False)
                    embed.set_footer(text="Created by The Invisible Man", icon_url="https://cdn.discordapp.com/avatars/366709133195476992/01cb7c2c7f2007d8b060e084ea4eb6fd.png?size=512")
                    await message.channel.send(embed=embed)
                    print("Completed")
                else:
                    messageSplit = message.content.split(" ")
                    channelNotif = messageSplit[1].split("#")
                    channelNotifSplit2 = channelNotif[1].split(">")
                    channelID = channelNotifSplit2[0]
                    channel = message.guild.get_channel(int(channelID))
                    async for mess in channel.history(limit=None):
                            messageTimeUTC = mess.created_at
                            source_time_zone = pytz.timezone("Etc/UTC")
                            source_date_with_timezone = source_time_zone.localize(messageTimeUTC)
                            target_time_zone = est
                            target_date_with_timezone = source_date_with_timezone.astimezone(target_time_zone)
                            hour = int(target_date_with_timezone.hour)
                            time[hour] = int(time[hour]) + 1
                    embed = discord.Embed(title="#" + channel.name + " Activity Times", color=0xFF9900)
                    for i in time:
                            embed.add_field(name=str(i) + ":00", value=str(time[i]), inline=False)
                    embed.set_footer(text="Created by The Invisible Man", icon_url="https://cdn.discordapp.com/avatars/366709133195476992/01cb7c2c7f2007d8b060e084ea4eb6fd.png?size=512")
                    await oneSec.delete()
                    await message.channel.send(embed=embed)
                    print("Completed")
            else:
                await message.channel.send("You do not have the required clearance to use this command.")
        else: 
            await message.channel.send("Sorry for the inconvience, but StatBot is currently shut down with most commands so people have the chance to opt in to the bot.")
    #Grabs the times info for each channel. 
    if message.content.startswith('&timeCountAll') and message.author.id == 366709133195476992:
        memberList = []
        for i in client.guilds: 
            print(i.name + "(" + str(i.id) + ")")
            for person in i.members:
                if person not in memberList and getUserOptLevel(person.id) >= 1:
                    memberList.append(person)
        print('Members Finished')
        number = 1
        wc = Workbook()
        wz = Workbook()
        sheet1 = wc.add_sheet('Sheet 1')
        sheet2 = wz.add_sheet('Sheet 1')
        sheet1.write(0, 1, "0:00")
        sheet1.write(0, 2, "1:00")
        sheet1.write(0, 3, "2:00")
        sheet1.write(0, 4, "3:00")
        sheet1.write(0, 5, "4:00")
        sheet1.write(0, 6, "5:00")
        sheet1.write(0, 7, "6:00")
        sheet1.write(0, 8, "7:00")
        sheet1.write(0, 9, "8:00")
        sheet1.write(0, 10, "9:00")
        sheet1.write(0, 11, "10:00")
        sheet1.write(0, 12, "11:00")
        sheet1.write(0, 13, "12:00")
        sheet1.write(0, 14, "13:00")
        sheet1.write(0, 15, "14:00")
        sheet1.write(0, 16, "15:00")
        sheet1.write(0, 17, "16:00")
        sheet1.write(0, 18, "17:00")
        sheet1.write(0, 19, "18:00")
        sheet1.write(0, 20, "19:00")
        sheet1.write(0, 21, "20:00")
        sheet1.write(0, 22, "21:00")
        sheet1.write(0, 23, "22:00")
        sheet1.write(0, 24, "23:00")

        sheet2.write(0, 1, "0:00")
        sheet2.write(0, 2, "1:00")
        sheet2.write(0, 3, "2:00")
        sheet2.write(0, 4, "3:00")
        sheet2.write(0, 5, "4:00")
        sheet2.write(0, 6, "5:00")
        sheet2.write(0, 7, "6:00")
        sheet2.write(0, 8, "7:00")
        sheet2.write(0, 9, "8:00")
        sheet2.write(0, 10, "9:00")
        sheet2.write(0, 11, "10:00")
        sheet2.write(0, 12, "11:00")
        sheet2.write(0, 13, "12:00")
        sheet2.write(0, 14, "13:00")
        sheet2.write(0, 15, "14:00")
        sheet2.write(0, 16, "15:00")
        sheet2.write(0, 17, "16:00")
        sheet2.write(0, 18, "17:00")
        sheet2.write(0, 19, "18:00")
        sheet2.write(0, 20, "19:00")
        sheet2.write(0, 21, "20:00")
        sheet2.write(0, 22, "21:00")
        sheet2.write(0, 23, "22:00")
        sheet2.write(0, 24, "23:00")

        wc.save("TimesMan.xls")
        wz.save("TimesManName.xls")
        authorDict = {}
        for beep in memberList:
            beepLevel = getUserOptLevel(beep.id)
            if getUserOptLevel(beep.id) > 0:
                print(beep.name + "(" + str(beep.id) + ")")
                sheet1.write(number, 0, str(beep.id))
                sheet2.write(number, 0, str(beep.name))
                number+=1
        wc.save("TimesMan.xls")
        wz.save("TimesManName.xls")
        timesTW = []
        
        for z in range(0, 24):
            newDict = {}
            for beep in memberList:
                if getUserOptLevel(beep.id) > 0:
                    newDict[beep.name] = 0
            timesTW.append(newDict)
        guildNumber = 1
        for i in client.guilds:
            channelNumber = 1
            for chan in i.text_channels:   
                if chan.name != "robot-game" and chan.name != "starfall-private-space" and chan.name != "ca-nerd-squad" and chan.name != "vent":
                    print(i.name + "; " + str(guildNumber) + "/" + str(len(client.guilds)) + ": " + chan.name + "; " + str(channelNumber) + "/" + str(len(i.text_channels))) 
                    async for mess in chan.history(limit=None):
                        if mess.author in memberList:
                            messageTimeUTC = mess.created_at
                            source_time_zone = pytz.timezone("Etc/UTC")
                            source_date_with_timezone = source_time_zone.localize(messageTimeUTC)
                            target_time_zone = est
                            target_date_with_timezone = source_date_with_timezone.astimezone(target_time_zone)
                            hourEST = int(target_date_with_timezone.hour)
                            correctDict = timesTW[hourEST - 1]
                            correctDict[mess.author.name] +=1
                    channelNumber+=1
            guildNumber+=1
        for i in range(len(timesTW)):
            for m in range(len(timesTW[i])):
                rowNum = m + 1
                curDict = list(timesTW[i].values())
                print("Value of Col " + str(i + 1) + ":" + str(curDict[m]))
                sheet1.write(rowNum, i + 1, curDict[m])
                sheet2.write(rowNum, i + 1, curDict[m])       
        wc.save("TimesMan.xls")
        wz.save("TimesManName.xls")
        dmChannel = await message.author.create_dm()
        await dmChannel.send("Times count complete.")
    #Counts the time for all users.
    if message.content.startswith('&userChart'):
        user = message.author
        userID = str(user.id)
        messageContentList = message.content.split(" ")
        if len(messageContentList) > 1:
            userID = 0
            if messageContentList[1].startswith("<@"): 
                userPing = messageContentList[1]
                userPingList = userPing.split('!')
                userPingThing = userPingList[1].split('>')
                userID = userPingThing[0]
            else:
                userID = messageContentList[1]
            user = client.get_user(int(userID))
        print("User Chart request made by " + message.author.name + " at " + str(message.created_at) + " on guild " + message.guild.name + " for " + user.name)
        if shutDown == False:
            if getUserOptLevel(userID) >= 2 or (getUserOptLevel >= 1 and message.author.id == userID):
                statBook = xlrd.open_workbook("C:\\Users\\Sebastian_Polge\\OneDrive-CaryAcademy\\Documents\\meNewBot\\Verity\\StatsBot\\TimesMan.xls")
                statSheet = statBook.sheet_by_index(0)
                userLoc = -1
                for i in range(statSheet.nrows):
                    if i != 0:
                        if str(user.id) == statSheet.cell_value(i, 0):
                            userLoc = i
                if userLoc == -1:
                    error()
                objects = ('0:00', '1:00', '2:00', '3:00', '4:00', '5:00', '6:00', '7:00', '8:00', '9:00', '10:00', '11:00', '12:00', '13:00', '14:00', '15:00', '16:00', '17:00', '18:00', '19:00', '20:00', '21:00', '22:00', '23:00')
                y_pos = np.arange(len(objects))
                topNum = 0
                for i in range(statSheet.ncols):
                    if i != 0:
                        if statSheet.cell_value(userLoc, i) > topNum:
                            topNum = statSheet.cell_value(userLoc, i)
                valueSplit = topNum / 5
                secNum = int(topNum - valueSplit)
                thirNum = int(secNum - valueSplit)
                fourNum = int(thirNum - valueSplit)
                fitNum = int(fourNum - valueSplit)
                per = [int(statSheet.cell_value(userLoc, 1)), int(statSheet.cell_value(userLoc, 2)), int(statSheet.cell_value(userLoc, 3)), int(statSheet.cell_value(userLoc, 4)), int(statSheet.cell_value(userLoc, 5)), int(statSheet.cell_value(userLoc, 6)), int(statSheet.cell_value(userLoc, 7)), int(statSheet.cell_value(userLoc, 8)), int(statSheet.cell_value(userLoc, 9)), int(statSheet.cell_value(userLoc, 10)), int(statSheet.cell_value(userLoc, 11)), int(statSheet.cell_value(userLoc, 12)), int(statSheet.cell_value(userLoc, 13)), int(statSheet.cell_value(userLoc, 14)), int(statSheet.cell_value(userLoc, 15)), int(statSheet.cell_value(userLoc, 16)), int(statSheet.cell_value(userLoc, 17)), int(statSheet.cell_value(userLoc, 18)), int(statSheet.cell_value(userLoc, 19)), int(statSheet.cell_value(userLoc, 20)), int(statSheet.cell_value(userLoc, 21)), int(statSheet.cell_value(userLoc, 22)), int(statSheet.cell_value(userLoc, 23)), int(statSheet.cell_value(userLoc, 24))]
                print(str(len(per)) + "/" + str(len(y_pos)))
                colorMember = message.guild.get_member(user.id)
                print(colorMember.name + ": " + str(colorMember.colour))
                fig = plt.figure(figsize=(75,50))
                plt.bar(y_pos, per, align='center', color=str(colorMember.colour))
                font = {'size'   : 75}

                plt.rc('font', **font)

                plt.xticks(y_pos, objects, fontsize=50)
                plt.yticks(fontsize=50)
                plt.ylabel('Messages')
                plt.title('Total Messages Sent All Time By Hour By ' + user.name)
                if user.name != ".| Sunny |.":
                    plt.savefig(user.name + '.png')
                else:
                    plt.savefig('Sunny.png')
                #plt.show()
                if user.name != ".| Sunny |.":
                    await message.channel.send(file=discord.File(user.name + '.png'))
                else:
                    await message.channel.send(file=discord.File('Sunny.png'))
            else:
                await message.channel.send("Your opt level is not high enough to access this information.")
        else:
            await message.channel.send("Sorry for the inconvience, but StatBot is currently shut down with most commands so people have the chance to opt in to the bot.")
    #Makes a chart of when a user sent the most messages. 
    if message.content.startswith('&fullChart') and message.author.id == 366709133195476992:
        statBook = xlrd.open_workbook("C:\\Users\\Sebastian_Polge\\OneDrive-CaryAcademy\\Documents\\meNewBot\\Verity\\StatsBot\\TimesManName.xls")
        statSheet = statBook.sheet_by_index(0)
        peopleStats = []
        peopleName = []
        for i in range(statSheet.nrows):
            if i != 0:
                perStat = []
                personName = statSheet.cell_value(i, 0)
                peopleName.append(personName)
                for locationNum in range(25):
                    if locationNum != 0:
                        quant = statSheet.cell_value(i, locationNum)
                        print("Quant of " + str(locationNum) + ": " + str(quant))
                        perStat.append(quant)
                peopleStats.append(perStat)
        bars = []
        for i in peopleStats:
            bar = i
            print("Bar: " + str(bar))
            bars.append(bar)
            

        # set width of bar
        barWidth = 0.25
        
        # set height of bar
        
        bars1 = [12, 30, 1, 8, 22, 12, 30, 1, 8, 22]
        bars2 = [28, 6, 16, 5, 10, 28, 6, 16, 5, 10]
        bars3 = [29, 3, 24, 25, 17, 29, 3, 24, 25, 17]
        
        # Set position of bar on X axis
        rList = []
        r1 = np.arange(24)
        lastR = r1
        timeThrough = 1
        for i in bars:
            R = [x + barWidth + 25 for x in lastR]
            rList.append(R)
            lastR = R

        # Make the plot
        import random
        color = []
        fig = plt.figure(figsize=(300,100))
        lastColor = "#557f2d"
        for i in range(len(bars)):
            print(rList[i])
            print(bars[i])
            if lastColor == "#557f2d":
                plt.bar(rList[i], bars[i], color='#7f6d5f', width=barWidth)
                lastColor = '#7f6d5f'
            else:
                plt.bar(rList[i], bars[i], color="#557f2d", width=barWidth)
                lastColor = "#557f2d"

        
        
        # Add xticks on the middle of the group bars
        plt.xlabel('People')
        plt.xticks([x + barWidth + 25 for x in lastR], peopleName)
        
        
        # Create legend & Show graphic
        plt.show()
    #My full chart command.
    if message.content.startswith('&fullChart') and message.author.id != 366709133195476992:
        await message.channel.send("Message `The Invisible Man#7937` for the full chart.")
    #When someone other than me wants to generate a full chart. 
    if message.content.startswith('&goodbyeToStatBot') and message.author.id == 366709133195476992:
        DnD = client.get_channel(474385207999463435)
        CADiscord = client.get_channel(524454096355459072)
        Level99 = client.get_channel(620788039911735332)
        Starserver = client.get_channel(621848134980730881)
        CoolPeeps = client.get_channel(364116594324013087)
        #await DnD.send("```After careful consideration and some complaints brought up by certain people about how StatBot collects data, I have decided to retire the bot, and delete all the associated data files. If anyone is particularly upset about the loss, thanks for your support, and send any messages to me at The Invisible Man#7937.``` -The Invisible Man.")
        #await DnD.send("Farewell to all, and to all a goodnight!")
        #await CADiscord.send("```After careful consideration and some complaints brought up by certain people about how StatBot collects data, I have decided to retire the bot, and delete all the associated data files. If anyone is particularly upset about the loss, thanks for your support, and send any messages to me at The Invisible Man#7937.``` -The Invisible Man.")
        #await CADiscord.send("Farewell to all, and to all a goodnight!")
        #await Level99.send("```After careful consideration and some complaints brought up by certain people about how StatBot collects data, I have decided to retire the bot, and delete all the associated data files. If anyone is particularly upset about the loss, thanks for your support, and send any messages to me at The Invisible Man#7937.``` -The Invisible Man.")
        #await Level99.send("Farewell to all, and to all a goodnight!")
        #await Starserver.send("```After careful consideration and some complaints brought up by certain people about how StatBot collects data, I have decided to retire the bot, and delete all the associated data files. If anyone is particularly upset about the loss, thanks for your support, and send any messages to me at The Invisible Man#7937.``` -The Invisible Man.")
        #await Starserver.send("Farewell to all, and to all a goodnight!")
        await CoolPeeps.send("```After careful consideration and some complaints brought up by certain people about how StatBot collects data, I have decided to retire the bot, and delete all the associated data files. If anyone is particularly upset about the loss, thanks for your support, and send any messages to me at The Invisible Man#7937.``` -The Invisible Man.")
        await CoolPeeps.send("Farewell to all, and to all a good night!")
        for i in client.guilds:
            if i.name != "test server":
                await i.leave()
    #Command used to remove StatBot from all servers, barring test server. 
    if message.content.startswith("&overTimeMessageChart") and message.author.id == 366709133195476992:        
        wa = Workbook()
        for server in client.guilds:
            userList = []
            anonUsers = []
            sept = {}
            octo = {}
            nove = {}
            dece = {}
            jan = {}
            feb = {}
            mar = {}
            apr = {}
            for member in server.members:
                if getUserOptLevel(member.id) >= 2 and member.bot == False:
                    print("Valid Member: " + member.name + " (" + str(member.id) + ")")
                    sept[member.name] = 0
                    octo[member.name] = 0
                    nove[member.name] = 0
                    dece[member.name] = 0
                    jan[member.name] = 0
                    feb[member.name] = 0
                    mar[member.name] = 0
                    apr[member.name] = 0
                    userList.append(member.name)
                if getUserOptLevel(member.id) >= 0:
                    print("Anon Member: " + member.name + " (" + str(member.id) + ")")
                    anonUsers.append(member.name)
            sept["anon"] = 0
            octo["anon"] = 0
            nove["anon"] = 0
            dece["anon"] = 0
            jan["anon"] = 0
            feb["anon"] = 0
            mar["anon"] = 0
            apr["anon"] = 0
            userList.append("anon")
            months = []
            month = 0
            months.append(sept)
            months.append(octo)
            months.append(nove)
            months.append(dece)
            months.append(jan)
            months.append(feb)
            months.append(mar)
            months.append(apr)

            for channel in server.text_channels:
                print("Checking #" + channel.name)
                newList = {}
                async for message in channel.history(limit=None):
                    if message.created_at.month == 9:
                        newList = months[0]
                    if message.created_at.month == 10:
                        newList = months[1]
                    if message.created_at.month == 11:
                        newList = months[2]
                    if message.created_at.month == 12:
                        newList = months[3]
                    if message.created_at.month == 1:
                        newList = months[4]
                    if message.created_at.month == 2 :
                        newList = months[5]
                    if message.created_at.month == 3:
                        newList = months[6]
                    if message.created_at.month == 4:
                        newList = months[7]
                    if message.author.name in userList:
                        newList[message.author.name]+=1
                    if message.author.name in anonUsers:
                        newList["anon"]+=1
                    
            print("Completed channels.")
            sheet1 = wa.add_sheet(server.name)
            i = 1
            for person in userList:
                sheet1.write(i, 0, person)
                i+=1
            sheet1.write(0, 1, "Sept.")
            sheet1.write(0, 2, "Oct.")
            sheet1.write(0, 3, "Nov.")
            sheet1.write(0, 4, "Dec.")
            sheet1.write(0, 5, "Jan.")
            sheet1.write(0, 6, "Feb.")
            sheet1.write(0, 7, "Mar.")
            sheet1.write(0, 8, "Apr.")

            zep = 0
            for moth in months:
                print("Anon Count: " + str(moth["anon"]))
                zeep = 1
                zep+=1
                for user in moth:
                    sheet1.write(zeep, zep, moth[user])
                    print(user + ": " + str(moth[user]))
                    zeep+=1
            wa.save("yes.xls")
            print("Completed!")
    #New command to make a chart. 
print("And we have hit the token. ")     
client.run('')